import os
import requests
from flask import Flask, render_template, request, redirect, url_for, jsonify, flash, session, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from PIL import Image
import stripe
from datetime import datetime, timedelta
from database import db
import uuid
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.barcode import code128
from reportlab.graphics import renderPDF
import io
import random
import string
import midtransclient
import json
import jwt
import sys # Import sys to check command line arguments

# Create the Flask app
app = Flask(__name__)

# Configuration
if not os.environ.get("SESSION_SECRET"):
    raise ValueError("SESSION_SECRET environment variable is required")
app.secret_key = os.environ.get("SESSION_SECRET")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL")
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Security configuration - Universal deployment ready
is_production = os.environ.get('IS_PRODUCTION', 'false').lower() == 'true'
debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'

app.config['SESSION_COOKIE_SECURE'] = is_production  # HTTPS only in production
app.config['SESSION_COOKIE_HTTPONLY'] = True  # No JS access
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['REMEMBER_COOKIE_SECURE'] = is_production
app.config['REMEMBER_COOKIE_HTTPONLY'] = True

# Initialize extensions
from database import configure_database
if not configure_database(app):
    print("[ERROR] Gagal mengkonfigurasi database, aplikasi akan berhenti")
    exit(1)

migrate = Migrate(app, db)
csrf = CSRFProtect(app)

# File upload configuration
app.config['UPLOAD_FOLDER'] = 'static/public/produk_images'
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max file size
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def compress_image(image_path, max_size_mb=1):
    """Compress image to be under max_size_mb while preserving original format"""
    img = Image.open(image_path)
    original_format = img.format

    # Check if file is already under size limit
    file_size_mb = os.path.getsize(image_path) / (1024 * 1024)
    if file_size_mb <= max_size_mb:
        return  # No compression needed

    # Don't compress animated GIFs - they lose animation
    if original_format == 'GIF' and getattr(img, 'is_animated', False):
        return  # Keep animated GIFs as-is

    # Set format-specific parameters
    save_kwargs = {'optimize': True}
    quality = 85

    if original_format in ('JPEG', 'JPG'):
        save_format = 'JPEG'
        save_kwargs['quality'] = quality
    elif original_format == 'PNG':
        save_format = 'PNG'
        # PNG doesn't use quality, but we can optimize
        save_kwargs['optimize'] = True
    elif original_format == 'WEBP':
        save_format = 'WEBP'
        save_kwargs['quality'] = quality
    else:
        # For other formats, convert to PNG to preserve quality
        save_format = 'PNG'
        save_kwargs['optimize'] = True

    # Try compression with decreasing quality for lossy formats
    max_iterations = 10
    iteration = 0

    while iteration < max_iterations:
        iteration += 1

        if save_format in ('JPEG', 'WEBP'):
            save_kwargs['quality'] = quality

        img.save(image_path, save_format, **save_kwargs)

        # Check file size
        file_size_mb = os.path.getsize(image_path) / (1024 * 1024)

        if file_size_mb <= max_size_mb:
            break

        # For lossy formats, try reducing quality
        if save_format in ('JPEG', 'WEBP'):
            if quality <= 20:
                break
            quality -= 10
        else:
            # For lossless formats (PNG), try downscaling dimensions
            current_width, current_height = img.size
            if current_width > 800 or current_height > 800:
                # Reduce by 80% each iteration
                new_width = int(current_width * 0.8)
                new_height = int(current_height * 0.8)
                img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            else:
                # Can't compress further, break to avoid infinite loop
                break

    return image_path

# Login Manager setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login untuk mengakses halaman ini.'

# Stripe configuration
if not os.environ.get('STRIPE_SECRET_KEY'):
    print("Warning: STRIPE_SECRET_KEY not set, using placeholder for development")
# Stripe API key will be set dynamically per request from PaymentConfiguration
# stripe.api_key = os.environ.get('STRIPE_SECRET_KEY', 'sk_test_placeholder_for_development')

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(models.User, int(user_id))

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Akses ditolak. Anda harus admin untuk mengakses halaman ini.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def staff_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or (not current_user.is_admin and not current_user.is_staff):
            flash('Akses ditolak. Anda harus staff atau admin untuk mengakses halaman ini.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# JWT Configuration for chat service integration
JWT_SECRET_KEY = os.environ.get("JWT_SECRET_KEY") or os.environ.get("SESSION_SECRET")  # Prefer separate JWT secret
JWT_ALGORITHM = 'HS256'
JWT_ACCESS_TOKEN_LIFETIME = 86400  # 24 hours

def generate_jwt_token(user):
    """Generate JWT token for chat service authentication"""
    payload = {
        'user_id': user.id,
        'email': user.email,
        'name': user.name,
        'role': user.role,
        'iat': datetime.utcnow(),
        'exp': datetime.utcnow() + timedelta(seconds=JWT_ACCESS_TOKEN_LIFETIME)
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

# Import models before routes
import models

def setup_django_chat_service():
    """Setup Django chat service and run migrations"""
    import subprocess
    import os
    import sys
    from pathlib import Path
    import time

    # Get absolute paths
    project_root = Path(__file__).resolve().parent
    chat_service_dir = project_root / 'chat_service'

    if not chat_service_dir.exists():
        print("[ERROR] Chat service directory not found")
        return False

    try:
        # Add paths to sys.path for Django imports
        sys.path.insert(0, str(chat_service_dir))
        sys.path.insert(0, str(project_root))

        # Set environment variables for Django
        os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'chat_microservice.settings')
        os.environ.setdefault('DJANGO_SECRET_KEY', app.secret_key)

        # Copy Flask database URL to Django
        if app.config.get('SQLALCHEMY_DATABASE_URI'):
            os.environ.setdefault('DATABASE_URL', app.config['SQLALCHEMY_DATABASE_URI'])
        
        # Set domains for Django ALLOWED_HOSTS
        os.environ.setdefault('DOMAINS', 'kasir.fajarmandiri.store,fajarmandiri.store')

        # Change to chat service directory
        original_cwd = os.getcwd()
        os.chdir(str(chat_service_dir))

        print("[SETUP] Running Django chat service migrations...")

        # Get Python executable path
        python_exec = sys.executable
        manage_py = chat_service_dir / 'manage.py'

        # Clean up existing migrations more safely
        migrations_dir = chat_service_dir / 'chat' / 'migrations'
        if migrations_dir.exists():
            for migration_file in migrations_dir.glob('*.py'):
                if migration_file.name != '__init__.py':
                    try:
                        migration_file.unlink()
                        print(f"[CLEANUP] Removed old migration: {migration_file.name}")
                    except Exception as e:
                        print(f"[WARNING] Could not remove {migration_file.name}: {e}")

        # Create __init__.py if it doesn't exist
        init_file = migrations_dir / '__init__.py'
        if not init_file.exists():
            migrations_dir.mkdir(parents=True, exist_ok=True)
            init_file.touch()

        # Make fresh migrations with better error handling
        makemigrations_cmd = [python_exec, str(manage_py), 'makemigrations', 'chat']
        result = subprocess.run(makemigrations_cmd, capture_output=True, text=True, timeout=30)

        if result.returncode == 0:
            print("[OK] Django migrations created")
        else:
            print(f"[WARNING] Makemigrations output: {result.stdout}")
            if result.stderr:
                print(f"[WARNING] Makemigrations stderr: {result.stderr}")

        # Apply migrations with better error handling
        migrate_cmd = [python_exec, str(manage_py), 'migrate', '--run-syncdb']
        result = subprocess.run(migrate_cmd, capture_output=True, text=True, timeout=60)

        if result.returncode == 0:
            print("[OK] Django chat service migrations completed")
        else:
            print(f"[WARNING] Migration output: {result.stdout}")
            if result.stderr:
                print(f"[WARNING] Migration stderr: {result.stderr}")
            # Try without --run-syncdb
            migrate_cmd = [python_exec, str(manage_py), 'migrate']
            result = subprocess.run(migrate_cmd, capture_output=True, text=True, timeout=60)
            if result.returncode == 0:
                print("[OK] Django chat service migrations completed (fallback)")

        # Test Django configuration
        test_cmd = [python_exec, '-c', '''
import os
import django
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "chat_microservice.settings")
django.setup()
from chat.models import ChatRoom
print("[TEST] Django models accessible")
''']
        result = subprocess.run(test_cmd, capture_output=True, text=True, timeout=15, cwd=str(chat_service_dir))

        if result.returncode == 0:
            print("[OK] Django chat service configured")
        else:
            print(f"[WARNING] Django test failed: {result.stderr}")

        # Start Django service if not already running
        try:
            import requests
            response = requests.get('http://127.0.0.1:8000/health/', timeout=2)
            if response.status_code == 200:
                print("[OK] Django chat service already running")
            else:
                print("[INFO] Starting Django chat service...")
                start_django_service()
        except:
            print("[INFO] Starting Django chat service...")
            start_django_service()

        # Return to original directory
        os.chdir(original_cwd)
        return True

    except subprocess.TimeoutExpired:
        print("[ERROR] Django setup timeout")
        os.chdir(original_cwd)
        return False
    except Exception as e:
        print(f"[ERROR] Django setup error: {e}")
        if 'original_cwd' in locals():
            os.chdir(original_cwd)
        return False

def start_django_service():
    """Start Django chat service in background"""
    import subprocess
    import sys
    from pathlib import Path

    try:
        project_root = Path(__file__).resolve().parent
        chat_service_dir = project_root / 'chat_service'
        manage_py = chat_service_dir / 'manage.py'

        # Start Django service in background
        subprocess.Popen([
            sys.executable, str(manage_py),
            'runserver', '0.0.0.0:8000', '--noreload'
        ], cwd=str(chat_service_dir))

        # Wait a moment for service to start
        import time
        time.sleep(3)

        print("[OK] Django chat service started on port 8000")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to start Django service: {e}")
        return False

def check_django_service():
    """Check if Django chat service is running"""
    try:
        import requests
        response = requests.get('http://127.0.0.1:8000/health/', timeout=2)
        return response.status_code == 200
    except:
        return False

def create_sample_data():
    """Create sample data for testing"""
    try:
        # Create categories
        categories_data = [
            {'name': 'Gitar', 'description': 'Gitar akustik dan elektrik'},
            {'name': 'Bass', 'description': 'Bass elektrik dan akustik'},
            {'name': 'Drum', 'description': 'Drum kit dan perkusi'},
            {'name': 'Keyboard', 'description': 'Keyboard dan piano digital'},
            {'name': 'Sound System', 'description': 'Speaker, mixer, dan audio equipment'},
        ]

        for cat_data in categories_data:
            existing = models.Category.query.filter_by(name=cat_data['name']).first()
            if not existing:
                category = models.Category(**cat_data)
                db.session.add(category)

        # Create suppliers
        suppliers_data = [
            {
                'name': 'Swelee Music Store',
                'contact_person': 'Swelee',
                'email': 'contact@swelee.com',
                'phone': '021-1234-5678',
                'company': 'PT Swelee Musik Indonesia'
            },
            {
                'name': 'Media Recording Tech',
                'contact_person': 'Media Recording',
                'email': 'info@mediarecording.com',
                'phone': '021-8765-4321',
                'company': 'CV Media Recording Technology'
            },
            {
                'name': 'Triple 3 Music Store',
                'contact_person': 'Triple Music',
                'email': 'sales@triple3music.com',
                'phone': '021-5555-3333',
                'company': 'Triple 3 Music Distribution'
            }
        ]

        for sup_data in suppliers_data:
            existing = models.Supplier.query.filter_by(name=sup_data['name']).first()
            if not existing:
                supplier = models.Supplier(**sup_data)
                db.session.add(supplier)

        # Create shipping services
        shipping_data = [
            {
                'name': 'JNE Regular',
                'code': 'jne_reg',
                'base_price': 15000,
                'price_per_kg': 5000,
                'min_days': 2,
                'max_days': 4
            },
            {
                'name': 'JNE Express',
                'code': 'jne_exp',
                'base_price': 25000,
                'price_per_kg': 8000,
                'min_days': 1,
                'max_days': 2
            },
            {
                'name': 'J&T Regular',
                'code': 'jnt_reg',
                'base_price': 12000,
                'price_per_kg': 4000,
                'min_days': 2,
                'max_days': 5
            }
        ]

        for ship_data in shipping_data:
            existing = models.ShippingService.query.filter_by(code=ship_data['code']).first()
            if not existing:
                service = models.ShippingService(**ship_data)
                db.session.add(service)

        db.session.commit()
        print("[OK] Sample data created successfully")

    except Exception as e:
        print(f"[ERROR] Error creating sample data: {e}")
        db.session.rollback()

# Global flag to prevent double initialization
_db_initialized = False

def initialize_database():
    """Initialize database with proper checks to prevent double initialization"""
    global _db_initialized

    if _db_initialized:
        print("[SKIP] Database already initialized")
        return

    try:
        # Ensure all Flask tables are created with current schema
        db.create_all()
        print("[OK] Flask database tables created")

        # Setup Django chat service
        if setup_django_chat_service():
            print("[OK] Django chat service setup completed")
        else:
            print("[WARNING] Django chat service setup failed, continuing without chat")

        # Create default admin user if it doesn't exist
        admin_email = "admin@hurtrock.com"

        # Check if admin user exists safely
        try:
            admin_user = models.User.query.filter_by(email=admin_email).first()
        except Exception as e:
            print(f"Database schema mismatch detected. Please run migration.")
            print(f"Error: {e}")
            admin_user = None

        if not admin_user:
            try:
                admin_user = models.User(
                    email=admin_email,
                    password_hash=generate_password_hash("admin123"),
                    name="Administrator",
                    role="admin"
                )
                db.session.add(admin_user)
                db.session.commit()
                print(f"[OK] Default admin user created: {admin_email}")
            except Exception as e:
                print(f"[ERROR] Failed to create admin user: {e}")
        else:
            print(f"[OK] Admin user already exists: {admin_email}")

        # Create default store profile if it doesn't exist
        try:
            store_profile = models.StoreProfile.get_active_profile()
            if not store_profile:
                store_profile = models.StoreProfile(
                    store_name='Hurtrock Music Store',
                    store_tagline='Toko Alat Musik Terpercaya',
                    store_address='Jl. Musik Raya No. 123, RT/RW 001/002, Kelurahan Musik, Kecamatan Harmoni',
                    store_city='Jakarta Selatan',
                    store_postal_code='12345',
                    store_phone='0821-1555-8035',
                    store_email='info@hurtrock.com',
                    store_website='https://hurtrock.com',
                    whatsapp_number='6282115558035',
                    operating_hours='Senin - Sabtu: 09:00 - 21:00\nMinggu: 10:00 - 18:00',
                    branch_name='Cabang Pusat',
                    branch_code='HRT-001'
                )
                db.session.add(store_profile)
                db.session.commit()
                print("[OK] Default store profile created")
            else:
                print("[OK] Store profile already exists")
        except Exception as e:
            print(f"[ERROR] Failed to create store profile: {e}")

        # Create sample data
        create_sample_data()

        # Mark as initialized
        _db_initialized = True
        print("[OK] Database initialization completed")

    except Exception as e:
        print(f"[ERROR] Database initialization error: {e}")

# Create database tables and setup everything
with app.app_context():
    initialize_database()

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html',
                         error_code=404,
                         error_message="Halaman yang Anda cari tidak ditemukan."), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('error.html',
                         error_code=500,
                         error_message="Terjadi kesalahan server. Silakan coba lagi nanti."), 500

# Routes
@app.route('/')
def index():
    products = models.Product.query.filter_by(is_active=True).limit(8).all()
    categories = models.Category.query.filter_by(is_active=True).all()
    return render_template('index.html', products=products, categories=categories)

@app.route('/products')
def products():
    category_id = request.args.get('category')
    search_query = request.args.get('search', '')

    query = models.Product.query.filter_by(is_active=True)

    if category_id:
        query = query.filter_by(category_id=category_id)

    if search_query:
        query = query.filter(models.Product.name.contains(search_query))

    products = query.all()
    categories = models.Category.query.filter_by(is_active=True).all()

    return render_template('products.html', products=products, categories=categories,
                         current_category=int(category_id) if category_id else None,
                         search_query=search_query)

@app.route('/product/<int:product_id>')
def product_detail(product_id):
    product = models.Product.query.get_or_404(product_id)
    return render_template('product_detail.html', product=product)

# This route was moved to avoid duplication - see line 1868-1884 for the actual implementation

@app.route('/search')
def search():
    query = request.args.get('q', '').strip()
    if query and len(query) >= 2:
        try:
            # Case-insensitive search across multiple fields (using ILIKE for PostgreSQL performance)
            search_term = f"%{query}%"
            products = models.Product.query.filter(
                db.or_(
                    models.Product.name.ilike(search_term),
                    models.Product.description.ilike(search_term),
                    models.Product.brand.ilike(search_term),
                    models.Product.model.ilike(search_term)
                ),
                models.Product.is_active == True
            ).limit(10).all()  # Limit results for performance

            return jsonify([{
                'id': p.id,
                'name': p.name,
                'price': str(p.price),
                'image_url': p.image_url or '/static/images/placeholder.jpg',
                'brand': p.brand or '',
                'description': p.description[:100] + '...' if p.description and len(p.description) > 100 else p.description or ''
            } for p in products])
        except Exception as e:
            print(f"Search error: {e}")
            return jsonify({'error': 'Search failed'}), 500
    return jsonify([])

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':

        email = request.form['email']
        password = request.form['password']
        name = request.form['name']

        # Check if user already exists
        if models.User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar. Silakan login.', 'error')
            return redirect(url_for('register'))

        # Create new user
        hashed_password = generate_password_hash(password)
        user = models.User(email=email, password_hash=hashed_password, name=name)

        db.session.add(user)
        db.session.commit()

        flash('Akun berhasil dibuat! Silakan login.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        user = models.User.query.filter_by(email=email).first()

        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            flash('Email atau password salah.', 'error')

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('index'))

@app.route('/cart')
@login_required
def cart():
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()
    total = sum(item.quantity * item.product.price for item in cart_items)
    return render_template('cart.html', cart_items=cart_items, total=total)

@app.route('/add_to_cart/<int:product_id>', methods=['POST'])
@login_required
def add_to_cart(product_id):
    product = models.Product.query.get_or_404(product_id)
    quantity = int(request.form.get('quantity', 1))

    # Check if item already in cart
    cart_item = models.CartItem.query.filter_by(
        user_id=current_user.id,
        product_id=product_id
    ).first()

    if cart_item:
        cart_item.quantity += quantity
    else:
        cart_item = models.CartItem(
            user_id=current_user.id,
            product_id=product_id,
            quantity=quantity
        )
        db.session.add(cart_item)

    db.session.commit()
    flash(f'{product.name} ditambahkan ke keranjang!', 'success')

    return redirect(url_for('product_detail', product_id=product_id))

@app.route('/update_cart/<int:item_id>', methods=['POST'])
@login_required
def update_cart(item_id):
    cart_item = models.CartItem.query.filter_by(
        id=item_id,
        user_id=current_user.id
    ).first_or_404()

    quantity = int(request.form.get('quantity', 1))

    if quantity > 0:
        cart_item.quantity = quantity
    else:
        db.session.delete(cart_item)

    db.session.commit()
    return redirect(url_for('cart'))

@app.route('/remove_from_cart/<int:item_id>')
@login_required
def remove_from_cart(item_id):
    cart_item = models.CartItem.query.filter_by(
        id=item_id,
        user_id=current_user.id
    ).first_or_404()

    db.session.delete(cart_item)
    db.session.commit()
    flash('Item dihapus dari keranjang.', 'info')

    return redirect(url_for('cart'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        current_user.name = request.form['name']
        current_user.phone = request.form['phone']
        current_user.address = request.form['address']

        db.session.commit()
        flash('Profile berhasil diperbarui!', 'success')

        # Redirect to checkout if came from there
        next_url = request.args.get('next')
        if next_url:
            return redirect(next_url)
        return redirect(url_for('profile'))

    return render_template('profile.html')

@app.route('/checkout')
@login_required
def checkout():
    # Check if buyer has complete profile (for buyers only)
    if current_user.is_buyer and (not current_user.address or not current_user.phone):
        flash('Silakan lengkapi profile Anda terlebih dahulu sebelum melakukan pembelian.', 'warning')
        return redirect(url_for('profile', next=url_for('checkout')))

    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if not cart_items:
        flash('Keranjang kosong!', 'error')
        return redirect(url_for('cart'))

    subtotal = sum(item.quantity * item.product.price for item in cart_items)

    # Calculate shipping weight and volume
    total_weight = sum(item.quantity * (item.product.weight or 0) for item in cart_items)
    total_volume = sum(item.quantity * (item.product.volume_cm3 or 0) for item in cart_items)

    # Get available shipping services
    shipping_services = models.ShippingService.query.filter_by(is_active=True).all()

    # Calculate shipping costs for each service
    shipping_options = []
    for service in shipping_services:
        cost = service.calculate_shipping_cost(total_weight, total_volume)
        shipping_options.append({
            'service': service,
            'cost': cost,
            'delivery_estimate': f"{service.min_days}-{service.max_days} hari"
        })

    # Get active payment configurations
    payment_configs = models.PaymentConfiguration.query.filter_by(is_active=True).all()

    # If no active payment config, show error
    if not payment_configs:
        flash('Tidak ada metode pembayaran yang tersedia. Silakan hubungi admin.', 'error')
        return redirect(url_for('cart'))

    return render_template('checkout.html',
                         cart_items=cart_items,
                         subtotal=subtotal,
                         total=subtotal,  # Add total variable (will be updated by JS when shipping is selected)
                         total_weight=total_weight,
                         total_volume=total_volume,
                         shipping_options=shipping_options,
                         payment_configs=payment_configs)

@app.route('/create-checkout-session', methods=['POST'])
@login_required
def create_checkout_session():
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if not cart_items:
        return jsonify({'error': 'Keranjang kosong'}), 400

    # Get selected payment configuration
    payment_config_id = request.form.get('payment_config_id')
    if not payment_config_id:
        flash('Silakan pilih metode pembayaran!', 'error')
        return redirect(url_for('checkout'))

    payment_config = models.PaymentConfiguration.query.get_or_404(int(payment_config_id))

    if not payment_config.is_active:
        flash('Metode pembayaran yang dipilih tidak aktif!', 'error')
        return redirect(url_for('checkout'))

    # Get selected shipping service
    shipping_service_id = request.form.get('shipping_service_id')
    if not shipping_service_id:
        flash('Silakan pilih jasa kirim!', 'error')
        return redirect(url_for('checkout'))

    shipping_service = models.ShippingService.query.get_or_404(int(shipping_service_id))

    # Calculate costs
    subtotal = sum(item.quantity * item.product.price for item in cart_items)
    total_weight = sum(item.quantity * (item.product.weight or 0) for item in cart_items)
    total_volume = sum(item.quantity * (item.product.volume_cm3 or 0) for item in cart_items)
    shipping_cost = shipping_service.calculate_shipping_cost(total_weight, total_volume)
    total_amount = float(subtotal) + float(shipping_cost)

    # Store order info in session
    session['shipping_service_id'] = shipping_service_id
    session['shipping_cost'] = float(shipping_cost)
    session['payment_config_id'] = payment_config_id

    try:
        YOUR_DOMAIN = os.environ.get('DOMAINS', 'localhost:5000')

        if payment_config.provider == 'stripe':
            return _create_stripe_checkout(cart_items, shipping_service, shipping_cost, total_amount, YOUR_DOMAIN, payment_config)
        elif payment_config.provider == 'midtrans':
            return _create_midtrans_checkout(cart_items, shipping_service, shipping_cost, total_amount, YOUR_DOMAIN, payment_config)
        else:
            flash('Metode pembayaran tidak didukung!', 'error')
            return redirect(url_for('checkout'))

    except Exception as e:
        flash(f'Error dalam memproses pembayaran: {str(e)}', 'error')
        return redirect(url_for('cart'))

def _create_stripe_checkout(cart_items, shipping_service, shipping_cost, total_amount, domain, payment_config):
    """Create Stripe checkout session"""

    # Set Stripe API key from config with fallback
    api_key = payment_config.stripe_secret_key or os.environ.get('STRIPE_SECRET_KEY')
    if not api_key:
        raise ValueError("No Stripe API key configured. Please configure payment settings in admin panel.")

    stripe.api_key = api_key

    # Build line items
    line_items = []
    for item in cart_items:
        line_items.append({
            'price_data': {
                'currency': 'idr',
                'product_data': {
                    'name': item.product.name,
                },
                'unit_amount': int(item.product.price * 100),  # Convert to cents
            },
            'quantity': item.quantity,
        })

    # Add shipping as line item
    line_items.append({
        'price_data': {
            'currency': 'idr',
            'product_data': {
                'name': f'Ongkos Kirim - {shipping_service.name}',
            },
            'unit_amount': int(shipping_cost * 100),  # Convert to cents
        },
        'quantity': 1,
    })

    checkout_session = stripe.checkout.Session.create(
        line_items=line_items,
        mode='payment',
        success_url=f'https://{domain}/payment-success',
        cancel_url=f'https://{domain}/cart',
        customer_email=current_user.email,
    )

    return redirect(checkout_session.url, code=303)

def _create_midtrans_checkout(cart_items, shipping_service, shipping_cost, total_amount, domain, payment_config):
    """Create Midtrans checkout session"""
    import uuid

    # Create Snap API instance
    snap = midtransclient.Snap(
        is_production=not payment_config.is_sandbox,
        server_key=payment_config.midtrans_server_key,
        client_key=payment_config.midtrans_client_key
    )

    # Generate unique order ID
    order_id = f"ORDER-{current_user.id}-{int(datetime.utcnow().timestamp())}-{str(uuid.uuid4())[:8]}"

    # Build item details
    item_details = []
    for item in cart_items:
        item_details.append({
            'id': str(item.product.id),
            'price': int(item.product.price),
            'quantity': item.quantity,
            'name': item.product.name[:50]  # Midtrans has character limit
        })

    # Add shipping cost
    item_details.append({
        'id': 'shipping',
        'price': int(shipping_cost),
        'quantity': 1,
        'name': f'Ongkir {shipping_service.name}'
    })

    # Customer details
    customer_details = {
        'first_name': current_user.name.split()[0] if current_user.name else 'Customer',
        'last_name': ' '.join(current_user.name.split()[1:]) if current_user.name and len(current_user.name.split()) > 1 else '',
        'email': current_user.email,
        'phone': current_user.phone or '081234567890',
        'billing_address': {
            'address': current_user.address or 'Jakarta',
            'city': 'Jakarta',
            'postal_code': '12345',
            'country_code': 'IDN'
        },
        'shipping_address': {
            'address': current_user.address or 'Jakarta',
            'city': 'Jakarta',
            'postal_code': '12345',
            'country_code': 'IDN'
        }
    }

    # Transaction details
    transaction_details = {
        'order_id': order_id,
        'gross_amount': int(total_amount)
    }

    # Parameter for Snap API
    param = {
        'transaction_details': transaction_details,
        'item_details': item_details,
        'customer_details': customer_details,
        'callbacks': {
            'finish': payment_config.callback_finish_url or f'https://{domain}/payment/finish',
            'unfinish': payment_config.callback_unfinish_url or f'https://{domain}/payment/unfinish',
            'error': payment_config.callback_error_url or f'https://{domain}/payment/error'
        }
    }

    # Create transaction
    transaction = snap.create_transaction(param)

    # Store order ID in session for callback handling
    session['midtrans_order_id'] = order_id

    # Redirect to Snap payment page
    return redirect(transaction['redirect_url'], code=303)

@app.route('/payment-success')
@login_required
def payment_success():
    # Create order from cart items
    cart_items = models.CartItem.query.filter_by(user_id=current_user.id).all()

    if cart_items:
        subtotal = sum(item.quantity * item.product.price for item in cart_items)

        # Get shipping info from session
        shipping_service_id = session.get('shipping_service_id')
        shipping_cost = session.get('shipping_cost', 0)

        total_amount = float(subtotal) + float(shipping_cost)

        # Calculate estimated delivery days
        estimated_delivery_days = 0
        if shipping_service_id:
            shipping_service = models.ShippingService.query.get(int(shipping_service_id))
            if shipping_service:
                estimated_delivery_days = shipping_service.max_days

        order = models.Order(
            user_id=current_user.id,
            total_amount=total_amount,
            shipping_cost=shipping_cost,
            shipping_service_id=int(shipping_service_id) if shipping_service_id else None,
            estimated_delivery_days=estimated_delivery_days,
            shipping_address=current_user.address,
            status='paid',
            created_at=datetime.utcnow()
        )
        db.session.add(order)
        db.session.flush()  # To get the order ID

        # Add order items
        for cart_item in cart_items:
            order_item = models.OrderItem(
                order_id=order.id,
                product_id=cart_item.product_id,
                quantity=cart_item.quantity,
                price=cart_item.product.price
            )
            db.session.add(order_item)

        # Clear cart
        for cart_item in cart_items:
            db.session.delete(cart_item)

        db.session.commit()

        flash('Pembayaran berhasil! Terima kasih atas pesanan Anda.', 'success')

    return render_template('payment_success.html', current_datetime=datetime.utcnow())

@app.route('/orders')
@login_required
def orders():
    user_orders = models.Order.query.filter_by(user_id=current_user.id).order_by(models.Order.created_at.desc()).all()
    return render_template('orders.html', orders=user_orders)

@app.route('/store-info')
def store_info():
    return render_template('store_info.html')





# Admin Routes
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    total_products = models.Product.query.count()
    total_orders = models.Order.query.count()
    total_users = models.User.query.count()

    recent_orders = models.Order.query.order_by(models.Order.created_at.desc()).limit(5).all()

    # Analisis penjualan
    from sqlalchemy import func, extract
    from decimal import Decimal

    # Total penjualan hari ini
    today = datetime.utcnow().date()
    today_sales = db.session.query(func.sum(models.Order.total_amount)).filter(
        func.date(models.Order.created_at) == today,
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).scalar() or Decimal('0')

    # Total penjualan bulan ini
    current_month = datetime.utcnow().month
    current_year = datetime.utcnow().year
    monthly_sales = db.session.query(func.sum(models.Order.total_amount)).filter(
        extract('month', models.Order.created_at) == current_month,
        extract('year', models.Order.created_at) == current_year,
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).scalar() or Decimal('0')

    # Produk terlaris dengan explicit join
    best_selling_products = db.session.query(
        models.Product.name,
        func.sum(models.OrderItem.quantity).label('total_sold')
    ).select_from(models.Product)\
    .join(models.OrderItem, models.OrderItem.product_id == models.Product.id)\
    .join(models.Order, models.Order.id == models.OrderItem.order_id)\
    .filter(
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).group_by(models.Product.id, models.Product.name).order_by(
        func.sum(models.OrderItem.quantity).desc()
    ).limit(5).all()

    return render_template('admin/dashboard.html',
                         total_products=total_products,
                         total_orders=total_orders,
                         total_users=total_users,
                         recent_orders=recent_orders,
                         current_date=datetime.utcnow(),
                         today_sales=today_sales,
                         monthly_sales=monthly_sales,
                         best_selling_products=best_selling_products)

@app.route('/admin/products')
@login_required
@admin_required
def admin_products():
    # Get all products ordered by stock status (critical first)
    products = models.Product.query.order_by(
        db.case(
            (models.Product.stock_quantity <= 0, 1),
            (models.Product.stock_quantity <= models.Product.minimum_stock, 2),
            (models.Product.stock_quantity <= models.Product.low_stock_threshold, 3),
            else_=4
        ),
        models.Product.name
    ).all()

    categories = models.Category.query.filter_by(is_active=True).all()
    suppliers = models.Supplier.query.filter_by(is_active=True).all()

    # Get products with different stock levels
    out_of_stock_products = models.Product.query.filter(models.Product.stock_quantity <= 0).all()
    critical_stock_products = models.Product.query.filter(
        models.Product.stock_quantity > 0,
        models.Product.stock_quantity <= models.Product.minimum_stock
    ).all()
    low_stock_products = models.Product.query.filter(
        models.Product.stock_quantity > models.Product.minimum_stock,
        models.Product.stock_quantity <= models.Product.low_stock_threshold
    ).all()

    # Combined alert products (critical + low stock)
    alert_products = critical_stock_products + low_stock_products

    return render_template('admin/products.html',
                         products=products,
                         categories=categories,
                         suppliers=suppliers,
                         out_of_stock_products=out_of_stock_products,
                         critical_stock_products=critical_stock_products,
                         low_stock_products=low_stock_products,
                         alert_products=alert_products)

@app.route('/admin/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_product():
    if request.method == 'GET':
        # Show add product form
        categories = models.Category.query.filter_by(is_active=True).all()
        suppliers = models.Supplier.query.filter_by(is_active=True).all()
        return render_template('admin/add_product.html', categories=categories, suppliers=suppliers)

    try:
        # Validate required fields
        if not request.form.get('name') or not request.form.get('price') or not request.form.get('category_id'):
            flash('Nama produk, harga, dan kategori wajib diisi!', 'error')
            categories = models.Category.query.filter_by(is_active=True).all()
            suppliers = models.Supplier.query.filter_by(is_active=True).all()
            return render_template('admin/add_product.html', categories=categories, suppliers=suppliers)

        name = request.form['name'].strip()
        description = request.form.get('description', '').strip()
        price = float(request.form['price'])
        stock_quantity = int(request.form.get('stock_quantity', 0))
        brand = request.form.get('brand', '').strip()
        model = request.form.get('model', '').strip()
        category_id = int(request.form['category_id'])
        supplier_id = int(request.form['supplier_id']) if request.form.get('supplier_id') else None
        is_featured = 'is_featured' in request.form
        is_active = 'is_active' in request.form

        # Handle dimensions and weight
        weight = float(request.form.get('weight', 0)) if request.form.get('weight') else 0
        length = float(request.form.get('length', 0)) if request.form.get('length') else 0
        width = float(request.form.get('width', 0)) if request.form.get('width') else 0
        height = float(request.form.get('height', 0)) if request.form.get('height') else 0

        # Handle stock management thresholds
        minimum_stock = int(request.form.get('minimum_stock', 5))
        low_stock_threshold = int(request.form.get('low_stock_threshold', 10))

        # Create new product
        new_product = models.Product(
            name=name,
            description=description,
            price=price,
            stock_quantity=stock_quantity,
            minimum_stock=minimum_stock,
            low_stock_threshold=low_stock_threshold,
            brand=brand,
            model=model,
            category_id=category_id,
            supplier_id=supplier_id,
            weight=weight,
            length=length,
            width=width,
            height=height,
            is_featured=is_featured,
            is_active=is_active
        )

        db.session.add(new_product)
        db.session.flush()  # Get product ID

        # Handle multiple images upload if provided
        uploaded_images = []
        selected_thumbnail_index = int(request.form.get('selected_thumbnail', 0))

        print(f"[DEBUG] Processing images for product {new_product.id}")
        print(f"[DEBUG] Selected thumbnail index: {selected_thumbnail_index}")

        if 'images' in request.files:
            files = request.files.getlist('images')
            print(f"[DEBUG] Found {len(files)} files in request")

            # Filter out empty files
            valid_files = [f for f in files if f and f.filename and f.filename.strip()]
            print(f"[DEBUG] {len(valid_files)} valid files to process")

            for i, file in enumerate(valid_files):
                if allowed_file(file.filename):
                    try:
                        # Generate unique filename
                        filename = secure_filename(file.filename)
                        filename = f"{uuid.uuid4()}_{filename}"
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                        # Create directory if it doesn't exist
                        os.makedirs(os.path.dirname(filepath), exist_ok=True)

                        # Save file
                        file.save(filepath)
                        print(f"[DEBUG] File saved: {filepath}")

                        # Compress image
                        compress_image(filepath)

                        image_url = f"/static/public/produk_images/{filename}"
                        uploaded_images.append(image_url)

                        # Create ProductImage record
                        is_thumbnail = (i == selected_thumbnail_index)
                        product_image = models.ProductImage(
                            product_id=new_product.id,
                            image_url=image_url,
                            is_thumbnail=is_thumbnail,
                            display_order=i
                        )
                        db.session.add(product_image)
                        print(f"[DEBUG] ProductImage created: {image_url}, is_thumbnail: {is_thumbnail}")

                        # Set the selected thumbnail as the main image_url
                        if is_thumbnail:
                            new_product.image_url = image_url
                            print(f"[DEBUG] Set main image_url: {image_url}")

                    except Exception as img_error:
                        print(f"[ERROR] Failed to process image {file.filename}: {str(img_error)}")
                        continue

        # If no thumbnail was selected but images were uploaded, use the first one
        if uploaded_images and not new_product.image_url:
            new_product.image_url = uploaded_images[0]
            # Update the first ProductImage to be thumbnail
            first_image = models.ProductImage.query.filter_by(product_id=new_product.id).first()
            if first_image:
                first_image.is_thumbnail = True
            print(f"[DEBUG] Using first image as thumbnail: {new_product.image_url}")

        db.session.commit()
        print(f"[DEBUG] Product {new_product.name} saved successfully with {len(uploaded_images)} images")
        flash(f'Produk {new_product.name} berhasil ditambahkan dengan {len(uploaded_images)} gambar!', 'success')

    except ValueError as ve:
        db.session.rollback()
        print(f"[ERROR] Validation error: {str(ve)}")
        flash(f'Data tidak valid: {str(ve)}', 'error')
        categories = models.Category.query.filter_by(is_active=True).all()
        suppliers = models.Supplier.query.filter_by(is_active=True).all()
        return render_template('admin/add_product.html', categories=categories, suppliers=suppliers)
    except Exception as e:
        db.session.rollback()
        print(f"[ERROR] Failed to add product: {str(e)}")
        flash(f'Gagal menambahkan produk: {str(e)}', 'error')

    return redirect(url_for('admin_products'))

@app.route('/admin/products/<int:product_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_product(product_id):
    product = models.Product.query.get_or_404(product_id)

    try:
        product.name = request.form['name']
        product.description = request.form.get('description', '')
        product.price = float(request.form['price'])
        product.stock_quantity = int(request.form['stock_quantity'])
        product.minimum_stock = int(request.form.get('minimum_stock', product.minimum_stock))
        product.low_stock_threshold = int(request.form.get('low_stock_threshold', product.low_stock_threshold))
        product.brand = request.form.get('brand', '')
        product.model = request.form.get('model', '')
        product.category_id = int(request.form['category_id'])
        product.supplier_id = int(request.form['supplier_id']) if request.form.get('supplier_id') else None
        product.is_featured = 'is_featured' in request.form
        product.is_active = 'is_active' in request.form

        # Handle dimensions and weight
        product.weight = float(request.form.get('weight', 0)) if request.form.get('weight') else 0
        product.length = float(request.form.get('length', 0)) if request.form.get('length') else 0
        product.width = float(request.form.get('width', 0)) if request.form.get('width') else 0
        product.height = float(request.form.get('height', 0)) if request.form.get('height') else 0

        # Handle multiple images upload if provided
        uploaded_images = []
        selected_thumbnail_index = int(request.form.get('selected_thumbnail', -1))

        # Check for new_images field (from edit form)
        files_to_process = []
        if 'new_images' in request.files:
            files_to_process = request.files.getlist('new_images')
        elif 'images' in request.files:
            files_to_process = request.files.getlist('images')

        if files_to_process:
            for i, file in enumerate(files_to_process):
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    filename = f"{uuid.uuid4()}_{filename}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                    # Create directory if it doesn't exist
                    os.makedirs(os.path.dirname(filepath), exist_ok=True)

                    file.save(filepath)
                    compress_image(filepath)

                    image_url = f"/static/public/produk_images/{filename}"
                    uploaded_images.append(image_url)

                    # Create ProductImage record
                    is_thumbnail = (i == selected_thumbnail_index)
                    product_image = models.ProductImage(
                        product_id=product.id,
                        image_url=image_url,
                        is_thumbnail=is_thumbnail,
                        display_order=len(product.images) + i  # Add after existing images
                    )
                    db.session.add(product_image)

                    # Set the selected thumbnail as the main image_url
                    if is_thumbnail:
                        product.image_url = image_url

        # Handle thumbnail selection from existing images
        current_thumbnail_id = request.form.get('current_thumbnail_selection')
        if current_thumbnail_id:
            # Update existing thumbnails
            for image in product.images:
                image.is_thumbnail = (str(image.id) == current_thumbnail_id)
                if image.is_thumbnail:
                    product.image_url = image.image_url

        # If new images were uploaded but no thumbnail selected, use first new image
        if uploaded_images and selected_thumbnail_index == -1:
            product.image_url = uploaded_images[0]
            # Find and update the first new image to be thumbnail
            newest_images = models.ProductImage.query.filter_by(product_id=product.id)\
                                                     .filter(models.ProductImage.image_url.in_(uploaded_images))\
                                                     .first()
            if newest_images:
                newest_images.is_thumbnail = True

        db.session.commit()
        flash(f'Produk {product.name} berhasil diperbarui!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_products'))

@app.route('/admin/products/<int:product_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_product(product_id):
    product = models.Product.query.get_or_404(product_id)

    # Check if product has been ordered
    if product.order_items:
        flash('Tidak bisa menghapus produk yang sudah pernah dipesan!', 'error')
        return redirect(url_for('admin_products'))

    # Check if product is in cart
    if product.cart_items:
        # Remove from all carts
        for cart_item in product.cart_items:
            db.session.delete(cart_item)

    product_name = product.name
    db.session.delete(product)
    db.session.commit()

    flash(f'Produk {product_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/product/<int:product_id>/edit', methods=['GET'])
@login_required
@admin_required
def admin_get_product(product_id):
    """Get product data for editing"""
    product = models.Product.query.get_or_404(product_id)

    return jsonify({
        'id': product.id,
        'name': product.name,
        'description': product.description or '',
        'price': float(product.price),
        'stock_quantity': product.stock_quantity,
        'minimum_stock': product.minimum_stock,
        'low_stock_threshold': product.low_stock_threshold,
        'brand': product.brand or '',
        'model': product.model or '',
        'category_id': product.category_id,
        'supplier_id': product.supplier_id,
        'is_active': product.is_active,
        'is_featured': product.is_featured,
        'image_url': product.image_url or '',
        'weight': float(product.weight or 0),
        'length': float(product.length or 0),
        'width': float(product.width or 0),
        'height': float(product.height or 0)
    })

@app.route('/admin/product/<int:product_id>/images', methods=['GET'])
@login_required
@admin_required
def admin_get_product_images(product_id):
    """Get product images for editing"""
    product = models.Product.query.get_or_404(product_id)

    # Get all product images ordered by display_order
    images = models.ProductImage.query.filter_by(product_id=product_id)\
                                      .order_by(models.ProductImage.display_order)\
                                      .all()

    image_data = []
    for image in images:
        image_data.append({
            'id': image.id,
            'image_url': image.image_url,
            'is_thumbnail': image.is_thumbnail,
            'display_order': image.display_order
        })

    return jsonify(image_data)

@app.route('/admin/categories/<int:category_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_category(category_id):
    category = models.Category.query.get_or_404(category_id)

    try:
        category.name = request.form['name']
        category.description = request.form.get('description', '')
        category.is_active = 'is_active' in request.form

        db.session.commit()
        flash(f'Kategori {category.name} berhasil diperbarui!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_categories'))

@app.route('/admin/categories/<int:category_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_category(category_id):
    category = models.Category.query.get_or_404(category_id)

    # Check if category has products
    if category.products:
        flash('Tidak bisa menghapus kategori yang memiliki produk!', 'error')
        return redirect(url_for('admin_categories'))

    category_name = category.name
    db.session.delete(category)
    db.session.commit()

    flash(f'Kategori {category_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_categories'))

@app.route('/admin/categories')
@login_required
@admin_required
def admin_categories():
    categories = models.Category.query.all()
    return render_template('admin/categories.html', categories=categories)

@app.route('/admin/categories/add', methods=['POST'])
@login_required
@admin_required
def admin_add_category():
    name = request.form['name']
    description = request.form['description']

    category = models.Category(name=name, description=description)
    db.session.add(category)
    db.session.commit()

    flash('Kategori berhasil ditambahkan!', 'success')
    return redirect(url_for('admin_categories'))


# Payment Configuration Routes
@app.route('/admin/payment-config')
@login_required
@admin_required
def admin_payment_config():
    configurations = models.PaymentConfiguration.query.all()
    return render_template('admin/payment_config.html', configurations=configurations)

@app.route('/admin/payment-config/create', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_create_payment_config():
    if request.method == 'POST':
        try:
            provider = request.form.get('provider')
            is_sandbox = request.form.get('is_sandbox') == 'true'

            # Validasi provider
            if not provider or provider not in ['midtrans', 'stripe']:
                flash('Provider pembayaran tidak valid!', 'error')
                return render_template('admin/create_payment_config.html')

            # Cek apakah sudah ada konfigurasi aktif untuk provider ini
            existing_config = models.PaymentConfiguration.query.filter_by(
                provider=provider,
                is_active=True
            ).first()

            if existing_config:
                flash(f'Sudah ada konfigurasi {provider} yang aktif! Nonaktifkan terlebih dahulu sebelum menambah yang baru.', 'warning')
                return render_template('admin/create_payment_config.html')

            config = models.PaymentConfiguration(
                provider=provider,
                is_sandbox=is_sandbox,
                is_active=False  # Start as inactive
            )

            if provider == 'midtrans':
                client_key = request.form.get('midtrans_client_key', '').strip()
                server_key = request.form.get('midtrans_server_key', '').strip()
                merchant_id = request.form.get('midtrans_merchant_id', '').strip()

                if not client_key or not server_key:
                    flash('Client Key dan Server Key harus diisi untuk Midtrans!', 'error')
                    return render_template('admin/create_payment_config.html')

                config.midtrans_client_key = client_key
                config.midtrans_server_key = server_key
                config.midtrans_merchant_id = merchant_id if merchant_id else None

            elif provider == 'stripe':
                publishable_key = request.form.get('stripe_publishable_key', '').strip()
                secret_key = request.form.get('stripe_secret_key', '').strip()

                if not publishable_key or not secret_key:
                    flash('Publishable Key dan Secret Key harus diisi untuk Stripe!', 'error')
                    return render_template('admin/create_payment_config.html')

                config.stripe_publishable_key = publishable_key
                config.stripe_secret_key = secret_key

            # Set callback URLs
            base_url = request.host_url.rstrip('/')
            config.callback_finish_url = f"{base_url}/payment/finish"
            config.callback_unfinish_url = f"{base_url}/payment/unfinish"
            config.callback_error_url = f"{base_url}/payment/error"
            config.notification_url = f"{base_url}/notification/handling"

            # Set additional URLs for Midtrans
            if provider == 'midtrans':
                config.recurring_notification_url = f"{base_url}/notification/recurring"
                config.account_linking_url = f"{base_url}/notification/account-linking"

            db.session.add(config)
            db.session.commit()

            environment_text = "Sandbox" if is_sandbox else "Production"
            flash(f'Konfigurasi pembayaran {provider.title()} ({environment_text}) berhasil ditambahkan!', 'success')
            return redirect(url_for('admin_payment_config'))

        except Exception as e:
            db.session.rollback()
            print(f"Error creating payment config: {str(e)}")
            flash(f'Terjadi kesalahan saat menyimpan konfigurasi: {str(e)}', 'error')
            return render_template('admin/create_payment_config.html')

    return render_template('admin/create_payment_config.html')

@app.route('/admin/payment-config/<int:config_id>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_payment_config(config_id):
    config = models.PaymentConfiguration.query.get_or_404(config_id)

    # Deactivate all other configs of the same provider
    if not config.is_active:
        models.PaymentConfiguration.query.filter_by(provider=config.provider).update({'is_active': False})

    config.is_active = not config.is_active
    config.updated_at = datetime.utcnow()

    db.session.commit()

    status = 'diaktifkan' if config.is_active else 'dinonaktifkan'
    flash(f'Konfigurasi {config.provider} berhasil {status}!', 'success')
    return redirect(url_for('admin_payment_config'))

# Midtrans Payment Callback Endpoints
@app.route('/payment/finish')
def payment_finish():
    order_id = request.args.get('order_id')
    status_code = request.args.get('status_code')
    transaction_status = request.args.get('transaction_status')

    if status_code == '200' and transaction_status == 'settlement':
        flash('Pembayaran berhasil! Terima kasih atas pesanan Anda.', 'success')
        return redirect(url_for('payment_success'))
    else:
        flash('Pembayaran gagal atau dibatalkan.', 'error')
        return redirect(url_for('cart'))

@app.route('/payment/unfinish')
def payment_unfinish():
    flash('Pembayaran belum diselesaikan. Silakan coba lagi.', 'warning')
    return redirect(url_for('cart'))

@app.route('/payment/error')
def payment_error():
    flash('Terjadi kesalahan dalam pembayaran. Silakan coba lagi.', 'error')
    return redirect(url_for('cart'))

@app.route('/payment/notification', methods=['POST'])
@csrf.exempt
def payment_notification():
    try:
        data = request.get_json()

        if not data:
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans',
            is_active=True
        ).first()

        if not midtrans_config:
            return jsonify({'status': 'error', 'message': 'Midtrans not configured'}), 400

        # Verify signature here (implementation depends on your requirements)
        order_id = data.get('order_id')
        transaction_status = data.get('transaction_status')
        fraud_status = data.get('fraud_status')

        if order_id:
            # Find the transaction
            midtrans_transaction = models.MidtransTransaction.query.filter_by(
                transaction_id=order_id
            ).first()

            if midtrans_transaction:
                midtrans_transaction.transaction_status = transaction_status
                midtrans_transaction.fraud_status = fraud_status
                midtrans_transaction.midtrans_response = json.dumps(data)
                midtrans_transaction.updated_at = datetime.utcnow()

                # Update order status based on transaction status
                if transaction_status == 'settlement' and fraud_status == 'accept':
                    midtrans_transaction.order.status = 'paid'
                elif transaction_status in ['deny', 'cancel', 'expire']:
                    midtrans_transaction.order.status = 'cancelled'

                db.session.commit()

        return jsonify({'status': 'ok'})

    except Exception as e:
        print(f"Payment notification error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal error'}), 500

@app.route('/notification/handling', methods=['POST'])
@csrf.exempt
def notification_handling():
    """
    URL Callback utama untuk notifikasi pembayaran Midtrans
    Menangani semua jenis notifikasi dari Midtrans
    """
    try:
        # Try to get JSON data, fallback to form data
        data = request.get_json()
        if not data:
            data = request.form.to_dict()

        if not data:
            print("No data received in notification")
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        # Log notifikasi untuk debugging
        print(f"Midtrans notification received: {json.dumps(data)}")

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans',
            is_active=True
        ).first()

        if not midtrans_config:
            print("No active Midtrans configuration found")
            # Return OK to prevent Midtrans from retrying
            return jsonify({'status': 'ok', 'message': 'No active Midtrans config'}), 200

        # Extract data from notification
        order_id = data.get('order_id', '')
        transaction_status = data.get('transaction_status', '')
        fraud_status = data.get('fraud_status', 'accept')
        payment_type = data.get('payment_type', '')
        gross_amount = data.get('gross_amount', '0')
        settlement_time = data.get('settlement_time')

        print(f"Processing order_id: {order_id}, status: {transaction_status}")

        if not order_id:
            print("No order_id in notification")
            return jsonify({'status': 'ok', 'message': 'No order_id provided'}), 200

        # Find or create the transaction record
        midtrans_transaction = models.MidtransTransaction.query.filter_by(
            transaction_id=order_id
        ).first()

        if not midtrans_transaction:
            print(f"Creating new transaction record for {order_id}")
            # Try to find order by various patterns
            order = None
            
            # Pattern 1: ORDER-{user_id}-{timestamp}-{uuid}
            try:
                order_parts = order_id.split('-')
                if len(order_parts) >= 3 and order_parts[0] == 'ORDER':
                    user_id = int(order_parts[1])
                    # Find recent unpaid order for this user
                    order = models.Order.query.filter_by(
                        user_id=user_id,
                        status='pending'
                    ).order_by(models.Order.created_at.desc()).first()
            except (ValueError, IndexError):
                pass

            # Pattern 2: Direct order ID
            if not order and order_id.isdigit():
                try:
                    order = models.Order.query.get(int(order_id))
                except ValueError:
                    pass

            # Pattern 3: Find by session data or recent orders
            if not order:
                # Find the most recent pending order
                order = models.Order.query.filter_by(
                    status='pending'
                ).order_by(models.Order.created_at.desc()).first()

            if order:
                midtrans_transaction = models.MidtransTransaction(
                    order_id=order.id,
                    transaction_id=order_id,
                    gross_amount=float(gross_amount) if gross_amount else 0,
                    payment_type=payment_type,
                    transaction_status=transaction_status,
                    fraud_status=fraud_status,
                    midtrans_response=json.dumps(data)
                )
                db.session.add(midtrans_transaction)
                db.session.flush()
                print(f"Created transaction record for order {order.id}")
            else:
                print(f"No matching order found for transaction {order_id}")
                return jsonify({'status': 'ok', 'message': 'No matching order found'}), 200

        if midtrans_transaction:
            # Update transaction details
            old_status = midtrans_transaction.transaction_status
            midtrans_transaction.transaction_status = transaction_status
            midtrans_transaction.fraud_status = fraud_status
            midtrans_transaction.payment_type = payment_type
            midtrans_transaction.midtrans_response = json.dumps(data)
            midtrans_transaction.updated_at = datetime.utcnow()

            if settlement_time:
                try:
                    midtrans_transaction.settlement_time = datetime.strptime(
                        settlement_time, '%Y-%m-%d %H:%M:%S'
                    )
                except ValueError:
                    print(f"Invalid settlement_time format: {settlement_time}")

            # Update order status based on transaction status
            order = midtrans_transaction.order
            old_order_status = order.status

            if transaction_status == 'settlement' and fraud_status == 'accept':
                order.status = 'paid'
                print(f"Order {order.id} marked as paid")
            elif transaction_status in ['deny', 'cancel', 'expire', 'failure']:
                order.status = 'cancelled'
                print(f"Order {order.id} marked as cancelled")
            elif transaction_status == 'pending':
                order.status = 'pending'
                print(f"Order {order.id} kept as pending")

            db.session.commit()

            print(f"Transaction {order_id} updated: {old_status} -> {transaction_status}, Order {order.id}: {old_order_status} -> {order.status}")

        return jsonify({'status': 'ok', 'message': 'Notification processed successfully'}), 200

    except Exception as e:
        print(f"Notification handling error: {str(e)}")
        import traceback
        traceback.print_exc()
        db.session.rollback()
        # Return OK to prevent Midtrans from retrying indefinitely
        return jsonify({'status': 'ok', 'message': 'Error processed, please check logs'}), 200

@app.route('/notification/recurring', methods=['POST'])
@csrf.exempt
def notification_recurring():
    """
    URL untuk notifikasi pembayaran berulang (subscription)
    """
    try:
        # Try to get JSON data, fallback to form data
        data = request.get_json()
        if not data:
            data = request.form.to_dict()

        if not data:
            print("No data received in recurring notification")
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        print(f"Recurring payment notification: {json.dumps(data)}")

        # Get active Midtrans configuration
        midtrans_config = models.PaymentConfiguration.query.filter_by(
            provider='midtrans',
            is_active=True
        ).first()

        if not midtrans_config:
            print("No active Midtrans configuration found for recurring")
            return jsonify({'status': 'ok', 'message': 'No active Midtrans config'}), 200

        # Extract recurring payment data
        subscription_id = data.get('subscription_id', '')
        transaction_id = data.get('transaction_id', '')
        transaction_status = data.get('transaction_status', '')
        payment_type = data.get('payment_type', '')
        gross_amount = data.get('gross_amount', '0')
        order_id = data.get('order_id', '')

        print(f"Processing recurring payment - subscription_id: {subscription_id}, transaction_id: {transaction_id}, status: {transaction_status}")

        # Log recurring payment attempt
        recurring_data = {
            'subscription_id': subscription_id,
            'transaction_id': transaction_id,
            'transaction_status': transaction_status,
            'payment_type': payment_type,
            'gross_amount': gross_amount,
            'order_id': order_id,
            'notification_data': data,
            'created_at': datetime.utcnow().isoformat()
        }

        # For now, we'll just log the recurring payment notification
        # In the future, this can be extended to handle subscription logic
        print(f"Recurring payment logged: {json.dumps(recurring_data)}")

        # If this is related to an existing order, try to update it
        if order_id:
            try:
                # Find transaction by order_id
                midtrans_transaction = models.MidtransTransaction.query.filter_by(
                    transaction_id=order_id
                ).first()

                if midtrans_transaction:
                    # Update with recurring payment info
                    response_data = json.loads(midtrans_transaction.midtrans_response or '{}')
                    response_data['recurring_info'] = data
                    midtrans_transaction.midtrans_response = json.dumps(response_data)
                    midtrans_transaction.updated_at = datetime.utcnow()
                    
                    db.session.commit()
                    print(f"Updated transaction {order_id} with recurring payment info")
                else:
                    print(f"No transaction found for recurring payment order_id: {order_id}")

            except Exception as e:
                print(f"Error updating transaction with recurring info: {str(e)}")

        # TODO: Implement subscription management logic here
        # This could include:
        # 1. Creating subscription records
        # 2. Managing recurring billing cycles
        # 3. Handling subscription renewals
        # 4. Sending renewal notifications to customers

        return jsonify({
            'status': 'ok',
            'message': 'Recurring payment notification processed',
            'subscription_id': subscription_id,
            'transaction_status': transaction_status
        }), 200

    except Exception as e:
        print(f"Recurring notification error: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return OK to prevent Midtrans from retrying indefinitely
        return jsonify({'status': 'ok', 'message': 'Error processed, please check logs'}), 200

@app.route('/notification/account-linking', methods=['POST'])
@csrf.exempt
def notification_account_linking():
    """
    URL untuk notifikasi menghubungkan akun (account linking)
    """
    try:
        data = request.get_json()

        if not data:
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        print(f"Account linking notification: {json.dumps(data)}")

        # Handle account linking logic here
        # This can be extended based on your account linking needs

        return jsonify({'status': 'ok'})

    except Exception as e:
        print(f"Account linking notification error: {str(e)}")
        return jsonify({'status': 'error', 'message': 'Internal error'}), 500

# API endpoint for cart count
@app.route('/api/cart/count')
@login_required
def api_cart_count():
    try:
        if current_user.role != 'buyer':
            return jsonify({'count': 0})

        count = models.CartItem.query.filter_by(user_id=current_user.id).count()
        return jsonify({'count': count})
    except Exception as e:
        print(f"Error getting cart count: {e}")
        return jsonify({'count': 0})

# API endpoint for chat service to get product info
@app.route('/api/products/<int:product_id>')
def api_get_product(product_id):
    try:
        product = models.Product.query.get_or_404(product_id)
        return jsonify({
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
            'image_url': product.image_url or '/static/images/placeholder.jpg',
            'brand': product.brand or '',
            'description': product.description or '',
            'category': product.category.name if product.category else '',
            'is_active': product.is_active
        })
    except Exception as e:
        print(f"Error getting product {product_id}: {e}")
        return jsonify({'error': 'Product not found'}), 404

# API endpoint for JWT token (for chat service)
@app.route('/api/chat/token')
@login_required
def api_chat_token():
    try:
        # Check if Django service is running
        if not check_django_service():
            print("[WARNING] Django chat service not responding, attempting to start...")
            start_django_service()
            
        token = generate_jwt_token(current_user)
        return jsonify({
            'token': token,
            'expires_in': JWT_ACCESS_TOKEN_LIFETIME,
            'user': {
                'id': current_user.id,
                'name': current_user.name,
                'email': current_user.email,
                'role': current_user.role
            }
        })
    except Exception as e:
        print(f"Error generating chat token: {str(e)}")
        return jsonify({'error': 'Failed to generate token'}), 500

# Proxy routes for chat service with /chat prefix
@app.route('/chat/<path:path>')
def proxy_chat_service(path):
    """Proxy all /chat requests to Django service on port 8000"""
    try:
        import requests
        
        # Build target URL
        target_url = f"http://127.0.0.1:8000/{path}"
        
        # Forward query parameters
        if request.query_string:
            target_url += f"?{request.query_string.decode()}"
        
        # Forward the request
        if request.method == 'GET':
            response = requests.get(target_url, headers=dict(request.headers), timeout=10)
        elif request.method == 'POST':
            response = requests.post(
                target_url,
                headers=dict(request.headers),
                data=request.get_data(),
                timeout=10
            )
        else:
            response = requests.request(
                request.method,
                target_url,
                headers=dict(request.headers),
                data=request.get_data(),
                timeout=10
            )
        
        # Return response
        return response.content, response.status_code, dict(response.headers)
        
    except Exception as e:
        print(f"Error proxying to chat service: {str(e)}")
        return jsonify({'error': 'Chat service unavailable'}), 503

# Chat service proxy endpoints
@app.route('/api/admin/buyer-rooms/')
@login_required
@admin_required
def proxy_buyer_rooms():
    try:
        # Check if Django service is running, if not try to start it
        if not check_django_service():
            print("[WARNING] Django chat service not responding, attempting to start...")
            if not start_django_service():
                return jsonify({'error': 'Chat service unavailable', 'rooms': [], 'total_count': 0}), 503

            # Wait for service to be ready
            import time
            for i in range(10):
                if check_django_service():
                    break
                time.sleep(1)
            else:
                return jsonify({'error': 'Chat service failed to start', 'rooms': [], 'total_count': 0}), 503

        search_query = request.args.get('search', '')

        # Try multiple endpoints for Django service
        endpoints = [
            'http://127.0.0.1:8000/api/admin/buyer-rooms/',
            'http://localhost:8000/api/admin/buyer-rooms/',
            'http://0.0.0.0:8000/api/admin/buyer-rooms/'
        ]

        if search_query:
            endpoints = [f"{url}?search={search_query}" for url in endpoints]

        headers = {
            'Authorization': f'Bearer {generate_jwt_token(current_user)}',
            'Content-Type': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                response = requests.get(chat_service_url, headers=headers, timeout=5)

                if response.status_code == 200:
                    return jsonify(response.json()), response.status_code
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({'error': 'Chat service unavailable', 'rooms': [], 'total_count': 0}), 503

    except Exception as e:
        print(f"Unexpected error in proxy_buyer_rooms: {str(e)}")
        return jsonify({'error': 'Internal server error', 'rooms': [], 'total_count': 0}), 500

@app.route('/api/rooms/<room_name>/messages/')
@login_required
def proxy_room_messages(room_name):
    try:
        # Check if Django service is running
        if not check_django_service():
            print("[WARNING] Django chat service not responding")
            return jsonify({'error': 'Chat service unavailable', 'results': []}), 503

        # Try multiple endpoints
        endpoints = [
            f"http://127.0.0.1:8000/api/rooms/{room_name}/messages/",
            f"http://localhost:8000/api/rooms/{room_name}/messages/",
            f"http://0.0.0.0:8000/api/rooms/{room_name}/messages/"
        ]

        # Forward query parameters
        query_params = request.args.to_dict()
        if query_params:
            query_string = "&".join([f"{k}={v}" for k, v in query_params.items()])
            endpoints = [f"{url}?{query_string}" for url in endpoints]

        headers = {
            'Authorization': f'Bearer {generate_jwt_token(current_user)}',
            'Content-Type': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                response = requests.get(chat_service_url, headers=headers, timeout=5)

                if response.status_code == 200:
                    try:
                        return jsonify(response.json()), response.status_code
                    except ValueError as e:
                        print(f"JSON decode error: {str(e)}")
                        return jsonify({'error': 'Invalid response from chat service', 'results': []}), 502
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({'error': 'Chat service unavailable', 'results': []}), 503

    except Exception as e:
        print(f"Unexpected error in proxy_room_messages: {str(e)}")
        return jsonify({'error': 'Internal server error', 'results': []}), 500

@app.route('/api/rooms/<room_name>/mark-read/', methods=['POST'])
@login_required
def proxy_mark_room_read(room_name):
    try:
        # Check if Django service is running
        if not check_django_service():
            print("[WARNING] Django chat service not responding")
            return jsonify({'error': 'Chat service unavailable'}), 503

        # Try multiple endpoints
        endpoints = [
            f"http://127.0.0.1:8000/api/rooms/{room_name}/mark-read/",
            f"http://localhost:8000/api/rooms/{room_name}/mark-read/",
            f"http://0.0.0.0:8000/api/rooms/{room_name}/mark-read/"
        ]

        headers = {
            'Authorization': f'Bearer {generate_jwt_token(current_user)}',
            'Content-Type': 'application/json'
        }

        last_error = None
        for chat_service_url in endpoints:
            try:
                response = requests.post(chat_service_url, headers=headers, timeout=5)

                if response.status_code == 200:
                    try:
                        return jsonify(response.json()), response.status_code
                    except ValueError:
                        return jsonify({'message': 'Messages marked as read'}), 200
                else:
                    last_error = f"Status {response.status_code}: {response.text}"
                    continue

            except requests.exceptions.RequestException as e:
                last_error = str(e)
                continue

        print(f"All Django endpoints failed. Last error: {last_error}")
        return jsonify({'error': 'Chat service unavailable'}), 503

    except Exception as e:
        print(f"Unexpected error in proxy_mark_room_read: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/admin/orders')
@login_required
@staff_required
def admin_orders():
    orders = models.Order.query.order_by(models.Order.created_at.desc()).all()
    return render_template('admin/orders.html', orders=orders)

@app.route('/admin/order/<int:order_id>/update', methods=['POST'])
@login_required
@staff_required
def admin_update_order(order_id):
    order = models.Order.query.get_or_404(order_id)

    new_status = request.form.get('status')
    tracking_number = request.form.get('tracking_number', '').strip()
    courier_service = request.form.get('courier_service', '').strip()

    if new_status not in ['pending', 'paid', 'shipped', 'delivered', 'cancelled']:
        flash('Status tidak valid!', 'error')
        return redirect(url_for('admin_orders'))

    order.status = new_status
    order.tracking_number = tracking_number if tracking_number else None
    order.courier_service = courier_service if courier_service else None
    order.updated_at = datetime.utcnow()

    db.session.commit()

    flash(f'Pesanan #{order.id} berhasil diperbarui!', 'success')
    return redirect(url_for('admin_orders'))

@app.route('/admin/order/<int:order_id>/quick-ship', methods=['POST'])
@login_required
@staff_required
def admin_quick_ship_order(order_id):
    order = models.Order.query.get_or_404(order_id)

    # Pastikan order masih dalam status paid
    if order.status != 'paid':
        flash(f'Pesanan #{order.id} tidak dalam status yang dapat dikirim!', 'error')
        return redirect(url_for('admin_orders'))

    tracking_number = request.form.get('tracking_number', '').strip()
    courier_service = request.form.get('courier_service', '').strip()
    
    if not tracking_number:
        tracking_number = generate_tracking_number()
    
    try:
        order.tracking_number = tracking_number
        order.courier_service = courier_service
        order.status = 'shipped'
        db.session.commit()
        
        flash(f'Pesanan #{order.id} berhasil dikirim dengan nomor resi: {tracking_number}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saat mengupdate status pengiriman: {str(e)}', 'error')
        
    return redirect(url_for('admin_orders'))

@app.route('/admin/order/<int:order_id>/print_professional_label')
@login_required
@staff_required
def print_professional_label(order_id):
    """
    Generate simple thermal label for 120mm printer
    Standard ReportLab format, clean and readable
    """
    order = models.Order.query.get_or_404(order_id)
    store_profile = models.StoreProfile.get_active_profile()

    # Generate tracking number if not exists
    if not order.tracking_number:
        order.tracking_number = generate_tracking_number()
        db.session.commit()

    buffer = io.BytesIO()

    # 120mm thermal printer width (340 points = 120mm)
    width = 340
    height = 480  # Adjustable height
    p = canvas.Canvas(buffer, pagesize=(width, height))

    y_pos = height - 20
    margin = 15

    # Header - Store Name
    p.setFont("Helvetica-Bold", 14)
    store_name = store_profile.store_name if store_profile else "Hurtrock Music Store"
    text_width = p.stringWidth(store_name, "Helvetica-Bold", 14)
    p.drawString((width - text_width) / 2, y_pos, store_name)
    y_pos -= 20

    # Separator line
    p.setLineWidth(1)
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # Order number
    p.setFont("Helvetica-Bold", 12)
    order_text = f"PESANAN #{order.id:06d}"
    text_width = p.stringWidth(order_text, "Helvetica-Bold", 12)
    p.drawString((width - text_width) / 2, y_pos, order_text)
    y_pos -= 15

    # Tracking number
    if order.tracking_number:
        p.setFont("Helvetica", 10)
        tracking_text = f"Resi: {order.tracking_number}"
        text_width = p.stringWidth(tracking_text, "Helvetica", 10)
        p.drawString((width - text_width) / 2, y_pos, tracking_text)
        y_pos -= 15

    # Date
    p.setFont("Helvetica", 9)
    date_text = f"Tanggal: {order.created_at.strftime('%d/%m/%Y %H:%M')}"
    text_width = p.stringWidth(date_text, "Helvetica", 9)
    p.drawString((width - text_width) / 2, y_pos, date_text)
    y_pos -= 20

    # Separator
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # DARI (Sender)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(margin, y_pos, "DARI:")
    y_pos -= 12

    p.setFont("Helvetica", 9)
    if store_profile:
        p.drawString(margin, y_pos, store_profile.store_name)
        y_pos -= 10

        # Split address into multiple lines
        address = store_profile.formatted_address
        words = address.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line + word) <= 35:
                current_line += word + " "
            else:
                if current_line:
                    lines.append(current_line.strip())
                current_line = word + " "
        if current_line:
            lines.append(current_line.strip())

        for line in lines[:3]:  # Max 3 lines
            p.drawString(margin, y_pos, line)
            y_pos -= 10

        if store_profile.store_phone:
            p.drawString(margin, y_pos, f"Telp: {store_profile.store_phone}")
            y_pos -= 10
    else:
        p.drawString(margin, y_pos, "Hurtrock Music Store")
        y_pos -= 10
        p.drawString(margin, y_pos, "Jakarta, Indonesia")
        y_pos -= 10

    y_pos -= 5

    # KEPADA (Recipient)
    p.setFont("Helvetica-Bold", 10)
    p.drawString(margin, y_pos, "KEPADA:")
    y_pos -= 12

    p.setFont("Helvetica-Bold", 9)
    p.drawString(margin, y_pos, order.user.name.upper())
    y_pos -= 12

    p.setFont("Helvetica", 8)
    if order.user.phone:
        p.drawString(margin, y_pos, f"Telp: {order.user.phone}")
        y_pos -= 10

    # Recipient address
    if order.user.address:
        address = order.user.address.replace('\n', ' ')
        words = address.split()
        lines = []
        current_line = ""
        for word in words:
            if len(current_line + word) <= 35:
                current_line += word + " "
            else:
                if current_line:
                    lines.append(current_line.strip())
                current_line = word + " "
        if current_line:
            lines.append(current_line.strip())

        for line in lines[:4]:  # Max 4 lines
            p.drawString(margin, y_pos, line)
            y_pos -= 10

    y_pos -= 5

    # Separator
    p.line(margin, y_pos, width - margin, y_pos)
    y_pos -= 15

    # Items
    p.setFont("Helvetica-Bold", 9)
    p.drawString(margin, y_pos, "BARANG:")
    y_pos -= 12

    p.setFont("Helvetica", 8)
    for item in order.order_items[:5]:  # Max 5 items
        item_name = item.product.name
        if len(item_name) > 30:
            item_name = item_name[:27] + "..."
        p.drawString(margin, y_pos, f"• {item_name}")
        y_pos -= 9
        p.drawString(margin + 10, y_pos, f"  {item.quantity}pcs")
        y_pos -= 10

    if len(order.order_items) > 5:
        p.drawString(margin, y_pos, f"• +{len(order.order_items) - 5} item lainnya")
        y_pos -= 10

    y_pos -= 5

    # Total and weight
    p.setFont("Helvetica-Bold", 10)
    total_text = f"TOTAL: {order.formatted_total}"
    text_width = p.stringWidth(total_text, "Helvetica-Bold", 10)
    p.drawString((width - text_width) / 2, y_pos, total_text)
    y_pos -= 15

    # Weight
    total_weight = sum(item.quantity * (item.product.weight or 100) for item in order.order_items) / 1000
    p.setFont("Helvetica", 9)
    weight_text = f"Berat: {total_weight:.1f} kg"
    text_width = p.stringWidth(weight_text, "Helvetica", 9)
    p.drawString((width - text_width) / 2, y_pos, weight_text)
    y_pos -= 15

    # Service info
    if order.courier_service:
        service_text = f"Kurir: {order.courier_service}"
        text_width = p.stringWidth(service_text, "Helvetica", 9)
        p.drawString((width - text_width) / 2, y_pos, service_text)
        y_pos -= 12

    # Footer
    p.setFont("Helvetica", 7)
    footer_text = "Terima kasih atas kepercayaan Anda"
    text_width = p.stringWidth(footer_text, "Helvetica", 7)
    p.drawString((width - text_width) / 2, 15, footer_text)

    p.showPage()
    p.save()

    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'thermal_label_{order.id}.pdf',
        mimetype='application/pdf'
    )

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = models.User.query.all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/analytics')
@login_required
@staff_required
def admin_analytics():
    from sqlalchemy import func, extract
    from decimal import Decimal

    # Penjualan 7 hari terakhir
    seven_days_ago = datetime.utcnow().date() - timedelta(days=6)
    daily_sales = db.session.query(
        func.date(models.Order.created_at).label('date'),
        func.sum(models.Order.total_amount).label('total'),
        func.count(models.Order.id).label('orders_count')
    ).filter(
        func.date(models.Order.created_at) >= seven_days_ago,
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).group_by(func.date(models.Order.created_at)).order_by(
        func.date(models.Order.created_at)
    ).all()

    # Penjualan per kategori dengan explicit join
    category_sales = db.session.query(
        models.Category.name,
        func.sum(models.OrderItem.quantity * models.OrderItem.price).label('total_sales'),
        func.sum(models.OrderItem.quantity).label('total_quantity')
    ).select_from(models.Category)\
    .join(models.Product, models.Product.category_id == models.Category.id)\
    .join(models.OrderItem, models.OrderItem.product_id == models.Product.id)\
    .join(models.Order, models.Order.id == models.OrderItem.order_id)\
    .filter(
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).group_by(models.Category.id, models.Category.name).order_by(
        func.sum(models.OrderItem.quantity * models.OrderItem.price).desc()
    ).all()

    # Pelanggan terbaik dengan explicit join
    top_customers = db.session.query(
        models.User.name,
        models.User.email,
        func.sum(models.Order.total_amount).label('total_spent'),
        func.count(models.Order.id).label('orders_count')
    ).select_from(models.User)\
    .join(models.Order, models.Order.user_id == models.User.id)\
    .filter(
        models.Order.status.in_(['paid', 'shipped', 'delivered'])
    ).group_by(models.User.id, models.User.name, models.User.email).order_by(
        func.sum(models.Order.total_amount).desc()
    ).limit(10).all()

    return render_template('admin/analytics.html',
                         daily_sales=daily_sales,
                         category_sales=category_sales,
                         top_customers=top_customers)

@app.route('/admin/user/<int:user_id>/change_role', methods=['POST'])
@login_required
@admin_required
def admin_change_user_role(user_id):
    user = models.User.query.get_or_404(user_id)
    new_role = request.form.get('role')

    if new_role not in ['admin', 'staff', 'buyer']:
        flash('Role tidak valid!', 'error')
        return redirect(url_for('admin_users'))

    # Prevent changing own role
    if user.id == current_user.id:
        flash('Tidak bisa mengubah role sendiri!', 'error')
        return redirect(url_for('admin_users'))

    user.role = new_role
    db.session.commit()

    flash(f'Role {user.name} berhasil diubah menjadi {new_role}!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_user(user_id):
    user = models.User.query.get_or_404(user_id)

    # Prevent deleting own account
    if user.id == current_user.id:
        flash('Tidak bisa menghapus akun sendiri!', 'error')
        return redirect(url_for('admin_users'))

    # Check if user has orders
    if user.orders:
        flash('Tidak bisa menghapus user yang memiliki riwayat pesanan!', 'error')
        return redirect(url_for('admin_users'))

    user_name = user.name
    db.session.delete(user)
    db.session.commit()

    flash(f'User {user_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/chat')
@login_required
@staff_required
def admin_chat():
    """Admin chat interface for customer service"""
    return render_template('admin/chat_interface.html')

@app.route('/chat')
@login_required
def chat():
    """Chat interface for buyers - placeholder page"""
    if current_user.is_admin or current_user.is_staff:
        return redirect(url_for('admin_chat'))

    # For now, redirect to a coming soon page or show message
    flash('Fitur chat sedang dalam pengembangan. Silakan hubungi admin melalui WhatsApp.', 'info')
    return redirect(url_for('index'))

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_user():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        role = request.form['role']

        # Validate role
        if role not in ['admin', 'staff', 'buyer']:
            flash('Role tidak valid!', 'error')
            return render_template('admin/add_user.html')

        # Check if email already exists
        if models.User.query.filter_by(email=email).first():
            flash('Email sudah terdaftar!', 'error')
            return render_template('admin/add_user.html')

        # Create new user
        hashed_password = generate_password_hash(password)
        user = models.User(
            name=name,
            email=email,
            password_hash=hashed_password,
            role=role
        )

        db.session.add(user)
        db.session.commit()

        flash(f'User {name} berhasil ditambahkan dengan role {role}!', 'success')
        return redirect(url_for('admin_users'))

    return render_template('admin/add_user.html')

@app.route('/admin/user/<int:user_id>/reset_password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(user_id):
    user = models.User.query.get_or_404(user_id)
    new_password = request.form.get('new_password')

    if not new_password or len(new_password) < 6:
        flash('Password harus minimal 6 karakter!', 'error')
        return redirect(url_for('admin_users'))

    user.password_hash = generate_password_hash(new_password)
    db.session.commit()

    flash(f'Password {user.name} berhasil direset!', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/export/sales/<period>')
@login_required
@staff_required
def export_sales(period):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Package openpyxl diperlukan untuk export Excel. Silakan install terlebih dahulu.', 'error')
        return redirect(url_for('admin_analytics'))

    from sqlalchemy import func
    import io
    import calendar

    # Get store profile
    store_profile = models.StoreProfile.get_active_profile()
    store_name = store_profile.store_name if store_profile else "Hurtrock Music Store"

    # Determine date range based on period
    today = datetime.utcnow().date()
    current_month = today.month
    current_year = today.year

    if period == 'daily':
        # Last 30 days
        start_date = today - timedelta(days=29)
        end_date = today
        period_text = f"Harian - {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}"

        sales_data = db.session.query(
            func.date(models.Order.created_at).label('period'),
            func.sum(models.Order.total_amount).label('total_sales'),
            func.count(models.Order.id).label('orders_count'),
            func.avg(models.Order.total_amount).label('avg_order_value')
        ).filter(
            func.date(models.Order.created_at) >= start_date,
            func.date(models.Order.created_at) <= end_date,
            models.Order.status.in_(['paid', 'shipped', 'delivered'])
        ).group_by(func.date(models.Order.created_at)).order_by(
            func.date(models.Order.created_at)
        ).all()

        date_format = lambda x: x.strftime('%d/%m/%Y')

    elif period == 'weekly':
        # Last 12 weeks
        start_date = today - timedelta(weeks=11)
        end_date = today
        period_text = f"Mingguan - {start_date.strftime('%d %B %Y')} s/d {end_date.strftime('%d %B %Y')}"

        sales_data = db.session.query(
            func.extract('year', models.Order.created_at).label('year'),
            func.extract('week', models.Order.created_at).label('week'),
            func.sum(models.Order.total_amount).label('total_sales'),
            func.count(models.Order.id).label('orders_count'),
            func.avg(models.Order.total_amount).label('avg_order_value')
        ).filter(
            func.date(models.Order.created_at) >= start_date,
            func.date(models.Order.created_at) <= end_date,
            models.Order.status.in_(['paid', 'shipped', 'delivered'])
        ).group_by(
            func.extract('year', models.Order.created_at),
            func.extract('week', models.Order.created_at)
        ).order_by(
            func.extract('year', models.Order.created_at),
            func.extract('week', models.Order.created_at)
        ).all()

        date_format = lambda x: f"Minggu {int(x.week)}/{int(x.year)}"

    elif period == 'monthly':
        # Last 12 months
        start_date = datetime(current_year - 1, current_month, 1).date()
        end_date = today
        period_text = f"Bulanan - {start_date.strftime('%B %Y')} s/d {end_date.strftime('%B %Y')}"

        sales_data = db.session.query(
            func.extract('year', models.Order.created_at).label('year'),
            func.extract('month', models.Order.created_at).label('month'),
            func.sum(models.Order.total_amount).label('total_sales'),
            func.count(models.Order.id).label('orders_count'),
            func.avg(models.Order.total_amount).label('avg_order_value')
        ).filter(
            func.date(models.Order.created_at) >= start_date,
            func.date(models.Order.created_at) <= end_date,
            models.Order.status.in_(['paid', 'shipped', 'delivered'])
        ).group_by(
            func.extract('year', models.Order.created_at),
            func.extract('month', models.Order.created_at)
        ).order_by(
            func.extract('year', models.Order.created_at),
            func.extract('month', models.Order.created_at)
        ).all()

        date_format = lambda x: f"{calendar.month_name[int(x.month)]} {int(x.year)}"

    else:
        flash('Periode tidak valid!', 'error')
        return redirect(url_for('admin_analytics'))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan Penjualan {period.title()}"

    # Set page orientation to landscape
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # Define styles
    header_font = Font(name='Arial', size=14, bold=True)
    title_font = Font(name='Arial', size=16, bold=True)
    subtitle_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    table_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')

    header_fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Company Header (Row 1-3)
    ws.merge_cells('A1:F1')
    company_cell = ws['A1']
    company_cell.value = store_name.upper()
    company_cell.font = title_font
    company_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    title_cell = ws['A2']
    title_cell.value = "Laporan Penjualan Sederhana"
    title_cell.font = subtitle_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A3:F3')
    period_cell = ws['A3']
    period_cell.value = f"Periode {period_text}"
    period_cell.font = subtitle_font
    period_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Empty row
    ws.row_dimensions[4].height = 10

    # Table headers (Row 5)
    headers = ['No', 'ID', 'Periode', 'Total Penjualan', 'Jumlah Pesanan', 'Rata-rata Order']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = table_header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    total_sales = 0
    total_orders = 0

    for row_idx, sale in enumerate(sales_data, 6):
        # No
        ws.cell(row=row_idx, column=1, value=row_idx - 5).border = border
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')

        # ID
        id_value = f"S{row_idx - 5:03d}"
        ws.cell(row=row_idx, column=2, value=id_value).border = border
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')

        # Period
        if period == 'daily':
            period_value = date_format(sale.period)
        else:
            period_value = date_format(sale)
        ws.cell(row=row_idx, column=3, value=period_value).border = border
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')

        # Total Sales
        sales_amount = float(sale.total_sales or 0)
        ws.cell(row=row_idx, column=4, value=sales_amount).border = border
        ws.cell(row=row_idx, column=4).number_format = '#,##0'
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='right')

        # Orders Count
        orders_count = sale.orders_count
        ws.cell(row=row_idx, column=5, value=orders_count).border = border
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center')

        # Average Order Value
        avg_value = float(sale.avg_order_value or 0)
        ws.cell(row=row_idx, column=6, value=avg_value).border = border
        ws.cell(row=row_idx, column=6).number_format = '#,##0'
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='right')

        total_sales += sales_amount
        total_orders += orders_count

    # Total row
    total_row = len(sales_data) + 6

    # Merge cells for "Total" label
    ws.merge_cells(f'A{total_row}:C{total_row}')
    total_label_cell = ws[f'A{total_row}']
    total_label_cell.value = "Total"
    total_label_cell.font = table_header_font
    total_label_cell.fill = header_fill
    total_label_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_label_cell.border = border

    # Total Sales
    total_sales_cell = ws.cell(row=total_row, column=4, value=total_sales)
    total_sales_cell.font = table_header_font
    total_sales_cell.fill = header_fill
    total_sales_cell.number_format = '#,##0'
    total_sales_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_sales_cell.border = border

    # Total Orders
    total_orders_cell = ws.cell(row=total_row, column=5, value=total_orders)
    total_orders_cell.font = table_header_font
    total_orders_cell.fill = header_fill
    total_orders_cell.alignment = Alignment(horizontal='center', vertical='center')
    total_orders_cell.border = border

    # Average (Total Sales / Total Orders)
    avg_total = total_sales / total_orders if total_orders > 0 else 0
    avg_total_cell = ws.cell(row=total_row, column=6, value=avg_total)
    avg_total_cell.font = table_header_font
    avg_total_cell.fill = header_fill
    avg_total_cell.number_format = '#,##0'
    avg_total_cell.alignment = Alignment(horizontal='right', vertical='center')
    avg_total_cell.border = border

    # Set column widths
    column_widths = [8, 12, 25, 20, 18, 20]  # Adjusted for A4 landscape
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Set row heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[5].height = 18

    # Footer
    footer_row = total_row + 2
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = f"Dicetak pada: {datetime.utcnow().strftime('%d %B %Y %H:%M:%S')}"
    footer_cell.font = Font(name='Arial', size=9, italic=True)
    footer_cell.alignment = Alignment(horizontal='center')

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'laporan_penjualan_{period}_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/admin/order/<int:order_id>/print_address')
@login_required
@staff_required
def print_order_address(order_id):
    order = models.Order.query.get_or_404(order_id)
    store_profile = models.StoreProfile.get_active_profile()

    # Create PDF untuk alamat pengiriman thermal (120mm width)
    buffer = io.BytesIO()

    # Calculate content
    content_lines = []

    # Header
    content_lines.extend([
        ('header', 'ALAMAT PENGIRIMAN'),
        ('divider', '=' * 35),
    ])

    # Store info (sender)
    content_lines.append(('section', 'DARI:'))
    if store_profile:
        content_lines.extend([
            ('store_name', store_profile.store_name),
            ('store_address', store_profile.formatted_address),
        ])
        if store_profile.store_phone:
            content_lines.append(('store_contact', f"Telp: {store_profile.store_phone}"))
        if store_profile.store_email:
            content_lines.append(('store_contact', f"Email: {store_profile.store_email}"))
    else:
        content_lines.extend([
            ('store_name', 'Hurtrock Music Store'),
            ('store_address', 'Jl. Musik Raya No. 123, Jakarta'),
            ('store_contact', 'Telp: 0821-1555-8035'),
        ])

    content_lines.append(('divider', '-' * 35))

    # Recipient info
    content_lines.extend([
        ('section', 'KEPADA:'),
        ('recipient_name', order.user.name.upper()),
    ])

    if order.user.phone:
        content_lines.append(('recipient_contact', f"Telp: {order.user.phone}"))

    if order.user.address:
        # Split long address
        address = order.user.address.replace('\n', ' ')
        if len(address) > 32:
            words = address.split(' ')
            lines = []
            current_line = ""
            for word in words:
                if len(current_line + word) <= 32:
                    current_line += word + " "
                else:
                    if current_line:
                        lines.append(current_line.strip())
                    current_line = word + " "
            if current_line:
                lines.append(current_line.strip())

            for line in lines:
                content_lines.append(('recipient_address', line))
        else:
            content_lines.append(('recipient_address', address))

    content_lines.append(('divider', '-' * 35))

    # Order info
    content_lines.extend([
        ('order_info', f"Order #{order.id}"),
        ('order_date', f"Tanggal: {order.created_at.strftime('%d/%m/%Y %H:%M')}"),
    ])

    if order.tracking_number:
        content_lines.append(('tracking', f"Resi: {order.tracking_number}"))

    if order.courier_service:
        content_lines.append(('courier', f"Kurir: {order.courier_service}"))

    content_lines.extend([
        ('total', f"Total: {order.formatted_total}"),
        ('divider', '=' * 35),
        ('footer', 'Terima kasih atas kepercayaan Anda'),
    ])

    # Calculate dimensions
    width = 340  # 120mm
    line_height = 12
    margin = 20
    total_height = margin * 2 + (len(content_lines) * line_height) + 40

    if total_height < 400:
        total_height = 400

    p = canvas.Canvas(buffer, pagesize=(width, total_height))

    y_pos = total_height - margin

    for line_type, text in content_lines:
        if line_type == 'header':
            p.setFont("Helvetica-Bold", 14)
            text_width = p.stringWidth(text, "Helvetica-Bold", 14)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 18
        elif line_type == 'divider':
            p.setFont("Helvetica", 10)
            text_width = p.stringWidth(text, "Helvetica", 10)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 12
        elif line_type == 'section':
            p.setFont("Helvetica-Bold", 12)
            p.drawString(margin, y_pos, text)
            y_pos -= 14
        elif line_type in ['store_name', 'recipient_name']:
            p.setFont("Helvetica-Bold", 11)
            p.drawString(margin + 5, y_pos, text)
            y_pos -= 13
        elif line_type in ['store_address', 'store_contact', 'recipient_address', 'recipient_contact']:
            p.setFont("Helvetica", 9)
            p.drawString(margin + 5, y_pos, text)
            y_pos -= 11
        elif line_type in ['order_info', 'order_date', 'tracking', 'courier']:
            p.setFont("Helvetica-Bold", 10)
            p.drawString(margin, y_pos, text)
            y_pos -= 12
        elif line_type == 'total':
            p.setFont("Helvetica-Bold", 12)
            text_width = p.stringWidth(text, "Helvetica-Bold", 12)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 15
        elif line_type == 'footer':
            p.setFont("Helvetica", 8)
            text_width = p.stringWidth(text, "Helvetica", 8)
            x_center = (width - text_width) / 2
            p.drawString(x_center, y_pos, text)
            y_pos -= 10
        else:
            p.setFont("Helvetica", 9)
            p.drawString(margin, y_pos, text)
            y_pos -= 11

    p.showPage()
    p.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'thermal_address_{order.id}.pdf',
        mimetype='application/pdf'
    )

# Restock Order Management
@app.route('/admin/restock')
@login_required
@admin_required
def admin_restock_orders():
    restock_orders = models.RestockOrder.query.order_by(models.RestockOrder.created_at.desc()).all()
    suppliers = models.Supplier.query.filter_by(is_active=True).all()
    products = models.Product.query.filter_by(is_active=True).all()

    # Convert to dict for JSON serialization
    restock_orders_data = []
    for order in restock_orders:
        order_data = {
            'id': order.id,
            'supplier': {
                'name': order.supplier.name,
                'contact_person': order.supplier.contact_person
            },
            'status': order.status,
            'total_amount': float(order.total_amount),
            'formatted_total': order.formatted_total,
            'created_at': order.created_at.isoformat(),
            'expected_date': order.expected_date.isoformat() if order.expected_date else None,
            'received_date': order.received_date.isoformat() if order.received_date else None,
            'notes': order.notes,
            'items': []
        }
        for item in order.items:
            order_data['items'].append({
                'product': {'name': item.product.name},
                'quantity_ordered': item.quantity_ordered,
                'unit_cost': float(item.unit_cost)
            })
        restock_orders_data.append(order_data)

    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'stock_quantity': product.stock_quantity
        })

    return render_template('admin/restock.html',
                         restock_orders=restock_orders,
                         suppliers=suppliers,
                         products=products,
                         restock_orders_data=restock_orders_data,
                         products_data=products_data)

@app.route('/admin/restock/create', methods=['POST'])
@login_required
@admin_required
def admin_create_restock_order():
    try:
        supplier_id = int(request.form['supplier_id'])
        notes = request.form.get('notes', '')
        expected_date_str = request.form.get('expected_date', '')

        expected_date = None
        if expected_date_str:
            expected_date = datetime.strptime(expected_date_str, '%Y-%m-%d')

        restock_order = models.RestockOrder(
            supplier_id=supplier_id,
            notes=notes,
            expected_date=expected_date,
            created_by=current_user.id
        )

        db.session.add(restock_order)
        db.session.flush()  # Get the order ID

        # Add items
        product_ids = request.form.getlist('product_ids[]')
        quantities = request.form.getlist('quantities[]')
        unit_costs = request.form.getlist('unit_costs[]')

        total_amount = 0
        for i, product_id in enumerate(product_ids):
            if product_id and quantities[i] and unit_costs[i]:
                quantity = int(quantities[i])
                unit_cost = float(unit_costs[i])

                item = models.RestockOrderItem(
                    restock_order_id=restock_order.id,
                    product_id=int(product_id),
                    quantity_ordered=quantity,
                    unit_cost=unit_cost
                )
                db.session.add(item)
                total_amount += quantity * unit_cost

        restock_order.total_amount = total_amount
        db.session.commit()

        flash(f'Order restock #{restock_order.id} berhasil dibuat!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_restock_orders'))

@app.route('/admin/restock/<int:order_id>/update_status', methods=['POST'])
@login_required
@admin_required
def admin_update_restock_status(order_id):
    restock_order = models.RestockOrder.query.get_or_404(order_id)

    try:
        new_status = request.form.get('status')
        if new_status not in ['pending', 'ordered', 'received', 'cancelled']:
            flash('Status tidak valid!', 'error')
            return redirect(url_for('admin_restock_orders'))

        restock_order.status = new_status

        if new_status == 'received':
            restock_order.received_date = datetime.utcnow()

            # Update product stock quantities
            for item in restock_order.items:
                item.quantity_received = item.quantity_ordered
                product = item.product
                product.stock_quantity += item.quantity_ordered

        db.session.commit()

        flash(f'Status restock order #{restock_order.id} berhasil diperbarui!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_restock_orders'))

@app.route('/admin/restock/<int:order_id>/invoice')
@login_required
@admin_required
def admin_generate_restock_invoice(order_id):
    restock_order = models.RestockOrder.query.get_or_404(order_id)

    # Create PDF invoice
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Header
    p.setFont("Helvetica-Bold", 18)
    p.drawString(50, height - 50, "INVOICE RESTOCK ORDER")

    # Store info
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 80, "DARI: Hurtrock Music Store")
    p.drawString(50, height - 95, "Jl. Musik Raya No. 123, Jakarta")
    p.drawString(50, height - 110, "Telp: 0821-1555-8035")

    # Invoice info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, height - 140, f"Invoice #: RESTOCK-{restock_order.id:05d}")
    p.drawString(50, height - 160, f"Tanggal: {restock_order.created_at.strftime('%d/%m/%Y')}")
    p.drawString(50, height - 180, f"Status: {restock_order.status.upper()}")

    # Supplier info
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, height - 140, "KEPADA:")
    p.setFont("Helvetica", 10)
    p.drawString(350, height - 160, f"{restock_order.supplier.name}")
    if restock_order.supplier.contact_person:
        p.drawString(350, height - 175, f"PIC: {restock_order.supplier.contact_person}")
    if restock_order.supplier.phone:
        p.drawString(350, height - 190, f"Telp: {restock_order.supplier.phone}")

    # Table header
    y_pos = height - 230
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_pos, "Produk")
    p.drawString(300, y_pos, "Qty")
    p.drawString(350, y_pos, "Harga Satuan")
    p.drawString(450, y_pos, "Total")

    # Draw line
    p.line(50, y_pos - 5, width - 50, y_pos - 5)

    # Items
    y_pos -= 25
    p.setFont("Helvetica", 9)
    total_amount = 0

    for item in restock_order.items:
        p.drawString(50, y_pos, item.product.name[:30])
        p.drawString(300, y_pos, str(item.quantity_ordered))
        p.drawString(350, y_pos, f"Rp {item.unit_cost:,.0f}".replace(',', '.'))
        p.drawString(450, y_pos, f"Rp {item.subtotal:,.0f}".replace(',', '.'))
        y_pos -= 15
        total_amount += item.subtotal

    # Total
    p.line(50, y_pos - 5, width - 50, y_pos - 5)
    y_pos -= 20
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, y_pos, "TOTAL:")
    p.drawString(450, y_pos, f"Rp {total_amount:,.0f}".replace(',', '.'))

    # Notes
    if restock_order.notes:
        y_pos -= 40
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y_pos, "Catatan:")
        y_pos -= 15
        p.setFont("Helvetica", 9)
        p.drawString(50, y_pos, restock_order.notes)

    p.showPage()
    p.save()

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f'invoice_restock_{restock_order.id:05d}.pdf',
        mimetype='application/pdf'
    )


# Shipping Services Management
@app.route('/admin/shipping')
@login_required
@admin_required
def admin_shipping_services():
    services = models.ShippingService.query.all()
    # Convert to dict for JSON serialization
    services_data = []
    for service in services:
        services_data.append({
            'id': service.id,
            'name': service.name,
            'code': service.code,
            'base_price': float(service.base_price),
            'price_per_kg': float(service.price_per_kg),
            'price_per_km': float(service.price_per_km),
            'volume_factor': float(service.volume_factor),
            'min_days': service.min_days,
            'max_days': service.max_days,
            'is_active': service.is_active
        })
    return render_template('admin/shipping.html', services=services, services_data=services_data)

@app.route('/admin/shipping/add', methods=['POST'])
@login_required
@admin_required
def admin_add_shipping_service():
    try:
        name = request.form['name']
        code = request.form['code']
        base_price = float(request.form['base_price'])
        price_per_kg = float(request.form['price_per_kg'])
        price_per_km = float(request.form.get('price_per_km', 0))
        volume_factor = float(request.form.get('volume_factor', 5000))
        min_days = int(request.form.get('min_days', 1))
        max_days = int(request.form.get('max_days', 3))

        service = models.ShippingService(
            name=name,
            code=code,
            base_price=base_price,
            price_per_kg=price_per_kg,
            price_per_km=price_per_km,
            volume_factor=volume_factor,
            min_days=min_days,
            max_days=max_days
        )

        db.session.add(service)
        db.session.commit()

        flash(f'Jasa kirim {name} berhasil ditambahkan!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_shipping_services'))

@app.route('/admin/shipping/<int:service_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_shipping_service(service_id):
    service = models.ShippingService.query.get_or_404(service_id)

    try:
        service.name = request.form['name']
        service.code = request.form['code']
        service.base_price = float(request.form['base_price'])
        service.price_per_kg = float(request.form['price_per_kg'])
        service.price_per_km = float(request.form.get('price_per_km', 0))
        service.volume_factor = float(request.form.get('volume_factor', 5000))
        service.min_days = int(request.form.get('min_days', 1))
        service.max_days = int(request.form.get('max_days', 3))
        service.is_active = request.form.get('is_active') == 'on'

        db.session.commit()

        flash(f'Jasa kirim {service.name} berhasil diperbarui!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_shipping_services'))

@app.route('/admin/shipping/<int:service_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_shipping_service(service_id):
    service = models.ShippingService.query.get_or_404(service_id)

    # Check if service is used in orders
    if service.orders:
        flash('Tidak bisa menghapus jasa kirim yang sedang digunakan di pesanan!', 'error')
        return redirect(url_for('admin_shipping_services'))

    service_name = service.name
    db.session.delete(service)
    db.session.commit()

    flash(f'Jasa kirim {service_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_shipping_services'))

# Supplier Management
@app.route('/admin/suppliers')
@login_required
@admin_required
def admin_suppliers():
    suppliers = models.Supplier.query.all()
    # Convert to dict for JSON serialization
    suppliers_data = []
    for supplier in suppliers:
        suppliers_data.append({
            'id': supplier.id,
            'name': supplier.name,
            'contact_person': supplier.contact_person,
            'email': supplier.email,
            'phone': supplier.phone,
            'address': supplier.address,
            'company': supplier.company,
            'notes': supplier.notes,
            'is_active': supplier.is_active
        })
    return render_template('admin/suppliers.html', suppliers=suppliers, suppliers_data=suppliers_data)

@app.route('/admin/suppliers/add', methods=['POST'])
@login_required
@admin_required
def admin_add_supplier():
    try:
        name = request.form['name']
        contact_person = request.form.get('contact_person', '')
        email = request.form.get('email', '')
        phone = request.form.get('phone', '')
        address = request.form.get('address', '')
        company = request.form.get('company', '')
        notes = request.form.get('notes', '')

        supplier = models.Supplier(
            name=name,
            contact_person=contact_person,
            email=email,
            phone=phone,
            address=address,
            company=company,
            notes=notes
        )

        db.session.add(supplier)
        db.session.commit()

        flash(f'Supplier {name} berhasil ditambahkan!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_suppliers'))

@app.route('/admin/suppliers/<int:supplier_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_supplier(supplier_id):
    supplier = models.Supplier.query.get_or_404(supplier_id)

    try:
        supplier.name = request.form['name']
        supplier.contact_person = request.form.get('contact_person', '')
        supplier.email = request.form.get('email', '')
        supplier.phone = request.form.get('phone', '')
        supplier.address = request.form.get('address', '')
        supplier.company = request.form.get('company', '')
        supplier.notes = request.form.get('notes', '')
        supplier.is_active = request.form.get('is_active') == 'on'

        db.session.commit()

        flash(f'Supplier {supplier.name} berhasil diperbarui!', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_suppliers'))

@app.route('/admin/suppliers/<int:supplier_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_delete_supplier(supplier_id):
    supplier = models.Supplier.query.get_or_404(supplier_id)

    # Check if supplier has products
    if supplier.products:
        flash('Tidak bisa menghapus supplier yang memiliki produk!', 'error')
        return redirect(url_for('admin_suppliers'))

    supplier_name = supplier.name
    db.session.delete(supplier)
    db.session.commit()

    flash(f'Supplier {supplier_name} berhasil dihapus!', 'success')
    return redirect(url_for('admin_suppliers'))

# Store Profile Management
@app.route('/admin/store-profile')
@login_required
@admin_required
def admin_store_profile():
    profile = models.StoreProfile.get_active_profile()
    if not profile:
        # Create default profile if none exists
        profile = models.StoreProfile(
            store_name='Hurtrock Music Store',
            store_tagline='Toko Alat Musik Terpercaya',
            store_address='Jl. Musik Raya No. 123',
            store_city='Jakarta',
            store_postal_code='12345',
            store_phone='0821-1555-8035',
            store_email='info@hurtrock.com'
        )
        db.session.add(profile)
        db.session.commit()

    return render_template('admin/store_profile.html', profile=profile)

@app.route('/admin/store-profile/update', methods=['POST'])
@login_required
@admin_required
def admin_update_store_profile():
    profile = models.StoreProfile.get_active_profile()
    if not profile:
        profile = models.StoreProfile()
        db.session.add(profile)

    try:
        # Basic store information
        profile.store_name = request.form.get('store_name', '').strip()
        profile.store_tagline = request.form.get('store_tagline', '').strip()
        profile.store_address = request.form.get('store_address', '').strip()
        profile.store_city = request.form.get('store_city', '').strip()
        profile.store_postal_code = request.form.get('store_postal_code', '').strip()
        profile.store_phone = request.form.get('store_phone', '').strip()
        profile.store_email = request.form.get('store_email', '').strip()
        profile.store_website = request.form.get('store_website', '').strip()

        # Branch information
        profile.branch_name = request.form.get('branch_name', '').strip()
        profile.branch_code = request.form.get('branch_code', '').strip()
        profile.branch_manager = request.form.get('branch_manager', '').strip()

        # Business information
        profile.business_license = request.form.get('business_license', '').strip()
        profile.tax_number = request.form.get('tax_number', '').strip()

        # Operating hours
        profile.operating_hours = request.form.get('operating_hours', '').strip()

        # Social media
        profile.facebook_url = request.form.get('facebook_url', '').strip()
        profile.instagram_url = request.form.get('instagram_url', '').strip()
        profile.whatsapp_number = request.form.get('whatsapp_number', '').strip()

        # Colors
        profile.primary_color = request.form.get('primary_color', '#FF6B35')
        profile.secondary_color = request.form.get('secondary_color', '#FF8C42')

        # Handle logo upload
        if 'logo' in request.files and request.files['logo'].filename:
            file = request.files['logo']
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filename = f"logo_{uuid.uuid4()}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                # Create directory if it doesn't exist
                os.makedirs(os.path.dirname(filepath), exist_ok=True)

                file.save(filepath)
                compress_image(filepath)
                profile.logo_url = f"/static/images/{filename}"

        profile.updated_at = datetime.utcnow()
        db.session.commit()

        flash('Profil toko berhasil diperbarui!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')

    return redirect(url_for('admin_store_profile'))

def generate_tracking_number():
    """Generate a tracking number for orders"""
    import random
    import string
    # Generate a simple tracking number format: TR + 8 random characters
    chars = string.ascii_uppercase + string.digits
    return 'TR' + ''.join(random.choices(chars, k=8))


if __name__ == '__main__':
    # Main execution
    if len(sys.argv) > 1 and sys.argv[1] == '--server-mode':
        # Running through server.py, don't start directly
        print("[INFO] Flask app initialized for server.py")
    else:
        # Direct execution
        print("[INFO] Starting Flask app directly...")
        print("[INFO] Main store will be available at: http://0.0.0.0:5000")

        # Check if running in production-like environment
        is_production = os.environ.get('IS_PRODUCTION', 'false').lower() == 'true'

        if is_production:
            # Production mode without reloader
            app.run(debug=False, host='0.0.0.0', port=5000, use_reloader=False)
        else:
            # Development mode without reloader for Replit compatibility  
            app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
